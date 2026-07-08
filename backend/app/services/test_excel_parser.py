import sys
import os

# Ensure the app module can be imported correctly
current_dir = os.path.dirname(os.path.abspath(__file__))
backend_dir = os.path.abspath(os.path.join(current_dir, "../../.."))
if backend_dir not in sys.path:
    sys.path.insert(0, backend_dir)

import openpyxl
from rapidfuzz import process, fuzz

from app.models.financial import FinancialRow, label_to_key
from app.services.excel_parser import _best_match, ANCHOR_KEYWORDS

def parse_messy_excel(file_path: str) -> list[FinancialRow]:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    # Strategy:
    # 1. Find the header row by looking for years (e.g., integers or strings like '2021', '2022')
    # 2. Assume the label is the first string encountered in a row.
    # 3. Associate values in the row with the columns that had year headers.

    years_col_map = {} # col_idx -> year_str
    
    # 1. Find years in the sheet
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        for col_idx, cell in enumerate(row):
            if cell is not None:
                cell_str = str(cell).strip()
                # Heuristic for year: 4 digits starting with 19 or 20
                if len(cell_str) == 4 and cell_str.isdigit() and cell_str.startswith(('19', '20')):
                    years_col_map[col_idx] = cell_str

        if years_col_map:
            break
            
    if not years_col_map:
        raise ValueError("Could not find year headers in the Excel file.")

    financial_rows = []
    order = 0

    # 2. Iterate rows, extract label and values
    for row in ws.iter_rows(values_only=True):
        label = None
        
        # Find the first string in the row to treat as label
        for cell in row:
            if cell is not None and isinstance(cell, str) and cell.strip():
                # Avoid treating the year headers as labels
                if cell.strip().isdigit() and len(cell.strip()) == 4:
                    continue
                # Avoid merged headers which might be very broad like "Fiscal Years"
                if cell.strip().lower() in ["fiscal years", "years", "line items", "metric"]:
                    continue
                label = cell.strip()
                break
                
        if not label:
            continue
            
        # Extract values for the years we found
        row_has_values = False
        year_values = {}
        for col_idx, year in years_col_map.items():
            try:
                val = row[col_idx]
                if val is not None:
                    # Clean up strings that might be numbers (e.g. "1,000.50")
                    if isinstance(val, str):
                        clean_val = val.replace(',', '').strip()
                        # handle negative values like (100) or -100
                        if clean_val.startswith('(') and clean_val.endswith(')'):
                            clean_val = '-' + clean_val[1:-1]
                        
                        try:
                            val = float(clean_val)
                        except ValueError:
                            val = None
                    elif isinstance(val, (int, float)):
                        val = float(val)
                    else:
                        val = None
                        
                    if val is not None:
                        year_values[year] = val
                        row_has_values = True
                    else:
                        year_values[year] = None
                else:
                    year_values[year] = None
            except (IndexError, TypeError, ValueError):
                year_values[year] = None

        if not row_has_values:
            continue # Probably a header or empty row
            
        # Match label
        matched_label, matched_section, statement_type, confidence = _best_match(label)
        meta = ANCHOR_KEYWORDS.get(matched_label, {})
        
        key = label_to_key(matched_label) if matched_label != "UNMAPPED" else label_to_key(label)

        frow = FinancialRow(
            row_id=f"{statement_type}_{order}",
            label=label,
            key=key,
            section=matched_section,
            is_subtotal=meta.get("is_subtotal", False),
            is_header=False,
            values=year_values,
            order=order
        )
        financial_rows.append(frow)
        order += 1

    return financial_rows

if __name__ == "__main__":
    dummy_path = os.path.join(current_dir, "dummy_messy.xlsx")
    if os.path.exists(dummy_path):
        print(f"Parsing {dummy_path}...")
        rows = parse_messy_excel(dummy_path)
        for r in rows:
            print(f"[{r.section}] {r.label} (Key: {r.key}, Confidence-based Match: {r.section}): {r.values}")
        print("Success!")
    else:
        print("Dummy file not found. Please run create_dummy.py first.")
