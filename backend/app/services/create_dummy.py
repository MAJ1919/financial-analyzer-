import openpyxl
from openpyxl.styles import Alignment, Font

def create_messy_excel(filepath):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Messy Financials"

    # Some blank rows at the top
    ws.append([])
    ws.append([])

    # Merged header for "Years" spanning columns B, C, D
    ws.merge_cells('B3:D3')
    ws['B3'] = 'Fiscal Years'
    ws['B3'].alignment = Alignment(horizontal='center')
    ws['B3'].font = Font(bold=True)

    # Actual headers in row 4, column A is empty or "Metric"
    ws['A4'] = 'Line Items'
    ws['B4'] = '2021'
    ws['C4'] = '2022'
    ws['D4'] = '2023'

    ws.append([]) # blank row

    # Data rows with some weird formatting or leading spaces
    # Revenue with a typo/alias
    ws.append(['  Total Revenue ', 1000, 1100, 1200])
    ws.append(['Cost of Goods Sold', 400, 450, 500])
    ws.append([]) # blank row
    # Subtotal
    ws.append(['Gross Profit', 600, 650, 700])
    
    # Another section with empty columns or something
    ws.append(['OpEx', 200, 220, 250])
    ws.append([' EBIT ', 400, 430, 450])

    ws.append([]) # blank row
    ws.append(['   Net Profit', 350, 380, 400])

    wb.save(filepath)
    print(f"Created {filepath}")

if __name__ == '__main__':
    create_messy_excel('dummy_messy.xlsx')
