"""
Excel Parser Service (Strict Template Mode)
============================================
Reads the raw .xlsx bytes.
STRICTLY expects sheets named exactly "Income Statement", "Balance Sheet", "Cash Flow Statement".
STRICTLY matches rows based on the exact labels provided in manualEntryTemplate.json.
If a user uploads an Excel file not matching this structure, it will fail gracefully.
"""

import difflib
import io
import re
import datetime as _dt
from typing import Any
import openpyxl
import math

from app.models.financial import FinancialRow, FinancialStatement, label_to_key
from app.models.statement_templates import load_statement_templates

# Canonical template — single source of truth in app/models/statement_templates.py
TEMPLATE_DATA = load_statement_templates()

# Strict sheet name mapping
STRICT_SHEET_MAP = {
    "income statement": "income_statement",
    "balance sheet": "balance_sheet",
    "cash flow statement": "cash_flow_statement",
}

# Labels THIS PROJECT shipped in earlier versions of Saudi_Template.xlsx.
#
# A curated exception to strict matching — and only that. Workbooks downloaded
# before the template was regenerated carry these labels, and the alternative
# remedy (download a fresh blank template) destroys whatever the user already
# entered. Matching stays exact: this maps one label the project is known to
# have published onto its current canonical label. It is NOT fuzzy matching,
# and unrecognised labels are still reported rather than guessed.
#
# Distinct from KEY_COMPAT_MAP in shared_utils.py, which remaps canonical
# *keys* for the analysis engines. This map is upload-time *labels* only.
#
# Keys are lowercased legacy labels; values are the current canonical label.
LEGACY_LABEL_ALIASES: dict[str, str] = {
    # Renamed when the download template was regenerated from the canonical JSON.
    "operating income": "Operating Income (EBIT)",
}

# A nearest-match below this ratio is too weak to suggest; the user is better
# served by "unmapped" alone than by a confidently wrong guess.
_SUGGESTION_CUTOFF = 0.6


def _clean_numeric_value(val: Any) -> float | None:
    if val is None:
        return None
    if isinstance(val, str):
        clean_val = val.replace(',', '').strip()
        if clean_val.startswith('(') and clean_val.endswith(')'):
            clean_val = '-' + clean_val[1:-1]
        try:
            parsed = float(clean_val)
            if math.isnan(parsed) or math.isinf(parsed):
                return None
            return parsed
        except ValueError:
            return None
    if isinstance(val, (int, float)):
        parsed = float(val)
        if math.isnan(parsed) or math.isinf(parsed):
            return None
        return parsed
    return None


def _extract_year_from_cell(cell: Any) -> str | None:
    """Strictly extract a 4-digit year from a cell."""
    if cell is None:
        return None
    if isinstance(cell, (_dt.datetime, _dt.date)):
        y = str(cell.year)
        return y if re.match(r"^(?:19|20)\d{2}$", y) else None
    # Numeric cells: Excel stores a typed "2022" as a number, not text
    if isinstance(cell, (int, float)) and not isinstance(cell, bool):
        if float(cell).is_integer():
            y = str(int(cell))
            return y if re.fullmatch(r"(?:19|20)\d{2}", y) else None
        return None
    if not isinstance(cell, str):
        return None
    val = cell.strip()
    if not val:
        return None
    # Strict full match for "2023", "FY2023", "2023A", etc.
    m = re.fullmatch(r"(?:FY)?((?:19|20)\d{2})[A-Za-z]?", val, re.IGNORECASE)
    if m:
        return m.group(1)
    return None


def parse_template_upload(file_bytes: bytes) -> dict:
    """
    Read .xlsx bytes, find global years, and STRICTLY map data row-by-row 
    to the JSON template labels. Non-matching rows are ignored.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    seen_year_ints: set[int] = set()
    global_years: list[str] = []

    # 1. Find global years by scanning all sheets for a valid header row
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                y = _extract_year_from_cell(cell)
                if y:
                    y_int = int(y[:4])
                    if y_int not in seen_year_ints:
                        seen_year_ints.add(y_int)
                        global_years.append(y)
            if global_years:
                break
        if global_years:
            break

    if not global_years:
        raise ValueError("Could not find valid year column headers (e.g., 2023, FY2023) in the Excel file. Please use the provided template.")

    global_years.sort(key=lambda y: int(y[:4]))
    result = {}

    # 2. Parse sheets strictly by template labels
    for sheet_name in wb.sheetnames:
        stmt_type = STRICT_SHEET_MAP.get(sheet_name.lower().strip())
        if not stmt_type:
            continue  # Skip sheets that aren't strictly named in our map

        template_rows = TEMPLATE_DATA.get(stmt_type, [])
        if not template_rows:
            continue

        ws = wb[sheet_name]
        
        # Find the header row (the row with the most matching global years)
        years_col_map: dict[int, str] = {}
        header_row_idx: int = 0

        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            temp_map: dict[int, str] = {}
            for col_idx, cell in enumerate(row):
                y = _extract_year_from_cell(cell)
                if y and y in global_years:
                    temp_map[col_idx] = y
            if len(temp_map) > len(years_col_map):
                years_col_map = temp_map
                header_row_idx = row_idx

        if not years_col_map:
            continue

        # Build a lookup dictionary of the Excel data: { "label_lowercase": [col1_val, col2_val...] }
        excel_data_lookup = {}
        duplicate_labels: list[str] = []
        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            if not row or not row[0] or not isinstance(row[0], str):
                continue
            label = row[0].strip()
            if not label:
                continue
            # First occurrence wins. A later duplicate used to overwrite it
            # silently, so one of the two rows vanished without ever being
            # reported — record it instead of dropping it.
            if label.lower() in excel_data_lookup:
                duplicate_labels.append(label)
                continue
            excel_data_lookup[label.lower()] = row

        # Redirect labels from templates this project previously shipped onto
        # their canonical row. An explicit canonical row in the sheet always
        # wins, so a partially-migrated workbook is never silently overridden.
        aliased_to: dict[str, str] = {}
        for legacy_lower, canonical in LEGACY_LABEL_ALIASES.items():
            canonical_lower = canonical.lower()
            if legacy_lower in excel_data_lookup and canonical_lower not in excel_data_lookup:
                excel_data_lookup[canonical_lower] = excel_data_lookup[legacy_lower]
                aliased_to[legacy_lower] = canonical_lower

        fin_rows = []
        used_labels = set()
        
        # Map STRICTLY by matching the template label to the Excel label
        for i, t_row in enumerate(template_rows):
            t_label = t_row.get("label", "").strip()
            excel_row = excel_data_lookup.get(t_label.lower())

            year_values = {}
            if excel_row:
                used_labels.add(t_label.lower())
                for target_year in global_years:
                    col_idx = next((c for c, y in years_col_map.items() if y == target_year), None)
                    if col_idx is not None and col_idx < len(excel_row):
                        try:
                            v = _clean_numeric_value(excel_row[col_idx])
                        except (IndexError, TypeError):
                            v = None
                    else:
                        v = None
                    year_values[target_year] = v
            else:
                # Label not found in Excel, populate with None
                year_values = {y: None for y in global_years}
                
            fin_rows.append(
                FinancialRow(
                    row_id=f"{stmt_type}_{i}",
                    label=t_row.get("label", t_label),
                    key=t_row.get("key", label_to_key(t_label)),
                    section=t_row.get("section", "Unknown"),
                    level=t_row.get("level", 3),
                    is_subtotal=t_row.get("is_subtotal", False),
                    is_header=t_row.get("is_header", False),
                    industry=t_row.get("industry", "general"),
                    values=year_values,
                    order=i,
                )
            )

        statement = FinancialStatement(years=global_years, rows=fin_rows)
        result[stmt_type] = statement.model_dump()

        # Track unmapped labels for this sheet, with a hint per label so the
        # user can act on the warning instead of diffing 236 canonical labels.
        canonical_labels = [r.get("label", "").strip() for r in template_rows if r.get("label")]
        unmapped = []
        hints: dict[str, str] = {}

        for label in excel_data_lookup:
            if label in used_labels:
                continue
            # A legacy label whose canonical row was matched is not unmapped —
            # its values were imported under the current label.
            if aliased_to.get(label) in used_labels:
                continue
            original_label = str(excel_data_lookup[label][0]).strip()
            if not original_label:
                continue
            unmapped.append(original_label)
            near = difflib.get_close_matches(
                original_label, canonical_labels, n=1, cutoff=_SUGGESTION_CUTOFF
            )
            if near:
                hints[original_label] = f'Did you mean "{near[0]}"?'

        for original_label in duplicate_labels:
            unmapped.append(original_label)
            hints[original_label] = "Duplicate row — only the first occurrence was imported."

        result.setdefault("unmapped_rows", {})
        result.setdefault("unmapped_hints", {})
        if unmapped:
            result["unmapped_rows"][sheet_name] = unmapped
        if hints:
            result["unmapped_hints"][sheet_name] = hints

    # 3. Ensure all 3 statements exist in the result using empty templates if missing
    # THIS FIXES THE MANUAL ENTRY BUG: Frontend always expects these keys to exist.
    for stmt_type in ["income_statement", "balance_sheet", "cash_flow_statement"]:
        if stmt_type not in result:
            fin_rows = []
            template_rows = TEMPLATE_DATA.get(stmt_type, [])
            for i, t_row in enumerate(template_rows):
                year_values = {y: None for y in global_years}
                fin_rows.append(
                    FinancialRow(
                        row_id=f"{stmt_type}_{i}",
                        label=t_row.get("label", ""),
                        key=t_row.get("key", ""),
                        section=t_row.get("section", "Unknown"),
                        level=t_row.get("level", 3),
                        is_subtotal=t_row.get("is_subtotal", False),
                        is_header=t_row.get("is_header", False),
                        industry=t_row.get("industry", "general"),
                        values=year_values,
                        order=i,
                    )
                )
            result[stmt_type] = FinancialStatement(years=global_years, rows=fin_rows).model_dump()

    # 4. Conditionally derive Cash Flow ONLY if it was completely missing/empty in the upload
    is_rows = {r["key"]: r for r in result["income_statement"]["rows"]}
    bs_rows = {r["key"]: r for r in result["balance_sheet"]["rows"]}
    cfs_rows = {r["key"]: r for r in result["cash_flow_statement"]["rows"]}

    # Check if the user actually provided CFS data in the Excel file
    has_uploaded_cfs_data = any(
        v is not None 
        for r in result["cash_flow_statement"]["rows"] 
        for v in r["values"].values()
    )

    if not has_uploaded_cfs_data and "income_statement" in result and "balance_sheet" in result:
        # Only run derivation if CFS was empty
        def get_is(key, year):
            if key in is_rows and year in is_rows[key]["values"] and is_rows[key]["values"][year] is not None:
                return float(is_rows[key]["values"][year])
            return 0.0

        def get_bs(key, year):
            if key in bs_rows and year in bs_rows[key]["values"] and bs_rows[key]["values"][year] is not None:
                return float(bs_rows[key]["values"][year])
            return 0.0
            
        def set_cfs(key, year, val):
            if key in cfs_rows:
                cfs_rows[key]["values"][year] = val

        for i in range(1, len(global_years)):
            year = global_years[i]
            prev_year = global_years[i - 1]

            ni = get_is("cfNetIncomeData", year)
            set_cfs("cfNetIncomeData", year, ni)
            
            da = abs(get_is("depreciationCostOfSales", year) + get_is("depreciationOpex", year))
            set_cfs("depreciationCostOfSales", year, get_is("depreciationCostOfSales", year)); set_cfs("depreciationOpex", year, get_is("depreciationOpex", year)); set_cfs("totalNonCashAdjustments", year, da)
            
            delta_ar = get_bs("tradeAccountsReceivable", prev_year) - get_bs("tradeAccountsReceivable", year)
            set_cfs("changeTradeAccountsReceivable", year, delta_ar)
            
            curr_inv = get_bs("rawMaterials", year) + get_bs("workInProcess", year) + get_bs("finishedGoods", year) + get_bs("otherInventory", year)
            prev_inv = get_bs("rawMaterials", prev_year) + get_bs("workInProcess", prev_year) + get_bs("finishedGoods", prev_year) + get_bs("otherInventory", prev_year)
            set_cfs("changeRawMaterials", year, prev_inv - curr_inv)
            
            delta_ap = get_bs("accountsPayable", year) - get_bs("accountsPayable", prev_year)
            set_cfs("changeAccountsPayable", year, delta_ap)
            
            curr_ppe = get_bs("grossPPE", year)
            prev_ppe = get_bs("grossPPE", prev_year)
            
            # FIX: CapEx derivation. Since PPE is Gross PPE, CapEx is just the delta.
            # Adding DA was double counting.
            capex = -(curr_ppe - prev_ppe)
            set_cfs("capitalExpenditures", year, capex)
            
            delta_st_debt = get_bs("stBorrowingsData", year) - get_bs("stBorrowingsData", prev_year)
            set_cfs("cfShortTermBorrowings", year, delta_st_debt)
            
            delta_lt_debt = get_bs("ltDebtData", year) - get_bs("ltDebtData", prev_year)
            set_cfs("cfLongTermBorrowings", year, delta_lt_debt)

    return result
