"""
Excel Export Service
====================
Builds a fully-formatted, formula-driven .xlsx workbook for a project.

Design (agreed with product owner):
  * Sheets: Income Statement, Balance Sheet, Cash Flow Statement, Ratios,
    Horizontal Analysis, DCF, Assumptions. The Assumptions sheet holds the
    editable driver panel on top AND, below it, the full forecast-engine
    calculation block (a live-formula port of forecasting_engine.py) in
    grouped rows, collapsed by default under a "do not edit" banner.
  * Historical columns hold the stored actuals (hardcoded inputs, blue).
  * Projected columns (…P) are a LIVE driver-based model: editing a driver on
    the Assumptions sheet (or a historical actual) recomputes every projection,
    exactly reproducing the app's ForecastingEngine (balanced/faithful modes).
  * Ratios, Horizontal Analysis and the DCF are live cross-sheet formulas.
  * Corporate/consulting theme: navy headers, thin rules, one light accent
    band over the projected year block. Blue = input, black = formula,
    green = cross-sheet link (standard financial-model colour convention).

The projected line-item DETAIL is scaled proportionally to revenue/assets and
the subtotals/totals are set to the engine aggregates — matching how
forecasting_engine.run_forecast() fills full_* statements, so the workbook
equals the Forecasting page for the same assumptions.
"""
from __future__ import annotations

import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from app.models.statement_templates import load_statement_templates
from app.models.financial import statement_to_lookup
from app.services.shared_utils import _parse_year
from app.services.forecasting_engine import calculate_historical_assumptions, extract_base_data
from dataclasses import asdict

# ============================================================
# THEME
# ============================================================
FONT_NAME = "Arial"
NAVY = "1F3864"
NAVY_MED = "2E4B7A"
GREY_HEADER = "D9E1F2"      # section header fill (light blue-grey)
ACCENT_PROJ = "EAF1FB"      # projected-year accent band
RULE = "BFBFBF"
WHITE = "FFFFFF"
BLUE_INPUT = "0000FF"       # hardcoded input
GREEN_LINK = "008000"       # cross-sheet link
BLACK = "000000"

FMT_NUM = '#,##0;(#,##0);"-"'
FMT_NUM1 = '#,##0.0;(#,##0.0);"-"'
FMT_PCT = '0.0%;(0.0%);"-"'
FMT_RATIO = '0.00"x";(0.00"x");"-"'
FMT_DAYS = '0.0;(0.0);"-"'

THIN = Side(style="thin", color=RULE)
MED = Side(style="medium", color=NAVY)

# The forecast-engine calculation block lives at the bottom of the
# Assumptions sheet (grouped rows under a "do not edit" banner); statement
# projections reference it there.
MODEL_SHEET = "Assumptions"


def _q(name: str) -> str:
    """Quote a sheet name for cross-sheet references."""
    return f"'{name}'"


def _font(bold=False, color=BLACK, size=10, italic=False):
    return Font(name=FONT_NAME, bold=bold, color=color, size=size, italic=italic)


# ============================================================
# LAYOUT PLANNING
# ============================================================
DATA_ROW0 = 6          # first line-item row on statement sheets
YEAR_COL0 = 3          # first year column (C); B = labels, A = margin
LABEL_COL = 2


class StmtLayout:
    """Coordinates for one statement sheet."""
    def __init__(self, sheet_name, title, template_rows, stored, hist_years, proj_years):
        self.sheet_name = sheet_name
        self.title = title
        self.template_rows = template_rows
        self.stored_lookup = statement_to_lookup(stored)
        self.hist_years = hist_years
        self.proj_years = proj_years
        self.years = hist_years + proj_years
        # year -> column letter
        self.year_col = {}
        for i, y in enumerate(self.years):
            self.year_col[y] = get_column_letter(YEAR_COL0 + i)
        self.base_year = hist_years[-1] if hist_years else None
        self.base_col = self.year_col.get(self.base_year)
        # key -> excel row  (row per template line item)
        self.key_row = {}
        for i, r in enumerate(template_rows):
            self.key_row[r["key"]] = DATA_ROW0 + i

    def cell(self, key, col):
        """Absolute-safe cell ref for a key at a given column letter, or '0'."""
        row = self.key_row.get(key)
        if row is None:
            return "0"
        return f"{_q(self.sheet_name)}!{col}{row}"

    def has(self, key):
        return key in self.key_row


# ============================================================
# ENGINE VARIABLE ORDER  (row per intermediate, cols = base + 5 forecast)
# ============================================================
ENGINE_VARS = [
    ("cogs_pct", "COGS % of revenue", FMT_PCT),
    ("sga_pct", "SG&A % of revenue", FMT_PCT),
    ("rd_pct", "R&D % of revenue", FMT_PCT),
    ("revenue", "Revenue", FMT_NUM),
    ("rev_mult", "Revenue multiplier", FMT_RATIO),
    ("ast_mult", "Asset multiplier", FMT_RATIO),
    ("cost_of_revenue", "Cost of Revenue", FMT_NUM),
    ("gross_profit", "Gross Profit", FMT_NUM),
    ("depreciation", "Depreciation", FMT_NUM),
    ("sga_expenses", "SG&A Expense", FMT_NUM),
    ("rd_expenses", "R&D Expense", FMT_NUM),
    ("total_op_expenses", "Total Operating Expenses", FMT_NUM),
    ("operating_income", "Operating Income (EBIT)", FMT_NUM),
    ("ebitda", "EBITDA", FMT_NUM),
    ("interest_income", "Interest Income", FMT_NUM),
    ("interest_expense", "Interest Expense", FMT_NUM),
    ("income_before_tax", "Income Before Tax", FMT_NUM),
    ("tax_expense", "Tax Expense", FMT_NUM),
    ("net_income", "Net Income", FMT_NUM),
    ("accounts_receivable", "Accounts Receivable", FMT_NUM),
    ("inventory", "Inventory", FMT_NUM),
    ("accounts_payable", "Accounts Payable", FMT_NUM),
    ("capex", "CapEx", FMT_NUM),
    ("cumulative_ppe", "Gross PPE", FMT_NUM),
    ("accumulated_depreciation", "Accumulated Depreciation", FMT_NUM),
    ("net_ppe", "Net PPE", FMT_NUM),
    ("wc_adjustment", "Working Capital Change", FMT_NUM),
    ("common_stock", "Common Stock", FMT_NUM),
    ("dividends_paid", "Dividends Paid", FMT_NUM),
    ("retained_earnings", "Retained Earnings", FMT_NUM),
    ("total_equity", "Total Equity", FMT_NUM),
    ("other_current_assets", "Other Current Assets", FMT_NUM),
    ("acquisitions", "Acquisitions", FMT_NUM),
    ("goodwill", "Goodwill / Other NCA", FMT_NUM),
    ("other_intangibles", "Other Intangibles", FMT_NUM),
    ("non_cash_assets", "Non-cash Assets", FMT_NUM),
    ("other_current_liabilities", "Other Current Liabilities", FMT_NUM),
    ("deferred_tax", "Deferred Tax", FMT_NUM),
    ("other_lt_liabilities", "Other LT Liabilities", FMT_NUM),
    ("long_term_debt", "Long-Term Debt", FMT_NUM),
    ("pre_revolver_liabilities", "Pre-revolver Liabilities", FMT_NUM),
    ("cash_surplus", "Cash Surplus", FMT_NUM),
    ("cash", "Cash", FMT_NUM),
    ("revolver", "Revolver", FMT_NUM),
    ("revolver_change", "Revolver Change", FMT_NUM),
    ("total_current_assets", "Total Current Assets", FMT_NUM),
    ("total_current_liabilities", "Total Current Liabilities", FMT_NUM),
    ("total_liabilities", "Total Liabilities", FMT_NUM),
    ("total_assets", "Total Assets", FMT_NUM),
    ("stock_based_comp", "Stock-based Comp", FMT_NUM),
    ("deferred_tax_change", "Deferred Tax Change", FMT_NUM),
    ("operating_cash_flow", "Operating Cash Flow", FMT_NUM),
    ("investing_cash_flow", "Investing Cash Flow", FMT_NUM),
    ("debt_issuance", "Debt Issuance", FMT_NUM),
    ("financing_cash_flow", "Financing Cash Flow", FMT_NUM),
    ("net_cash_change", "Net Change in Cash", FMT_NUM),
    ("free_cash_flow", "Free Cash Flow", FMT_NUM),
]
# Engine rows are assigned dynamically below the Assumptions driver panel
# (see build_workbook) — banner row, block header row, then one row per var.


# ============================================================
# STATEMENT PROJECTION OVERRIDES  (template key -> f(E, colL) formula body)
# E = quoted engine sheet ref prefix helper; colL = engine forecast column letter
# Each returns the formula WITHOUT the leading '=' .
# ============================================================
def _build_overrides(engine_row):
    """engine_row: dict name->row number on the engine sheet."""
    E = _q(MODEL_SHEET)

    def e(name, col):
        return f"{E}!{col}{engine_row[name]}"

    is_ov = {
        "revenueHeader": lambda c: e("revenue", c),
        "totalRevenue": lambda c: e("revenue", c),
        "costOfRevenueDisplayHeader": lambda c: e("cost_of_revenue", c),
        "totalCostOfRevenue": lambda c: e("cost_of_revenue", c),
        "grossProfit": lambda c: e("gross_profit", c),
        "grossProfitHeader": lambda c: e("gross_profit", c),
        "operatingExpensesHeader": lambda c: f"-{e('total_op_expenses', c)}",
        "totalSellingExpense": lambda c: f"-{e('sga_expenses', c)}*0.5",
        "totalGeneralAdminExpense": lambda c: f"-{e('sga_expenses', c)}",
        "researchAndDevelopment": lambda c: f"-{e('rd_expenses', c)}",
        "depreciationOpex": lambda c: f"-{e('depreciation', c)}",
        "operatingIncomeDisplayHeader": lambda c: e("operating_income", c),
        "operatingIncome": lambda c: e("operating_income", c),
        "nonOperatingHeader": lambda c: f"{e('income_before_tax', c)}-{e('operating_income', c)}",
        "financeCosts": lambda c: f"-{e('interest_expense', c)}",
        "incomeBeforeTax": lambda c: e("income_before_tax", c),
        "earningsBeforeTax": lambda c: e("income_before_tax", c),
        "incomeTaxExpense": lambda c: f"-{e('tax_expense', c)}",
        "currentIncomeTax": lambda c: f"-{e('tax_expense', c)}",
        "netIncome": lambda c: e("net_income", c),
        "netIncomeAttributableToParent": lambda c: e("net_income", c),
        "totalComprehensiveIncome": lambda c: e("net_income", c),
        "ebitda": lambda c: e("ebitda", c),
    }
    bs_ov = {
        "cashAndEquivalents": lambda c: e("cash", c),
        "receivablesHeader": lambda c: e("accounts_receivable", c),
        "netReceivables": lambda c: e("accounts_receivable", c),
        "inventoryHeader": lambda c: e("inventory", c),
        "totalInventory": lambda c: e("inventory", c),
        "currentAssetsHeader": lambda c: e("total_current_assets", c),
        "totalCurrentAssets": lambda c: e("total_current_assets", c),
        "ppeHeader": lambda c: e("net_ppe", c),
        "netPPE": lambda c: e("net_ppe", c),
        "grossPPE": lambda c: e("cumulative_ppe", c),
        "accumulatedDepreciation": lambda c: e("accumulated_depreciation", c),
        "intangibleAssetsHeader": lambda c: f"{e('goodwill', c)}+{e('other_intangibles', c)}",
        "grossIntangibleAssets": lambda c: f"{e('goodwill', c)}+{e('other_intangibles', c)}",
        "nonCurrentAssetsHeader": lambda c: f"{e('net_ppe', c)}+{e('goodwill', c)}+{e('other_intangibles', c)}",
        "totalNonCurrentAssets": lambda c: f"{e('net_ppe', c)}+{e('goodwill', c)}+{e('other_intangibles', c)}",
        "assetsHeader": lambda c: e("total_assets", c),
        "totalAssets": lambda c: e("total_assets", c),
        "stBorrowingsData": lambda c: e("revolver", c),
        "currentPortionLTDebt": lambda c: "0",
        "currentLiabilitiesHeader": lambda c: e("total_current_liabilities", c),
        "totalCurrentLiabilities": lambda c: e("total_current_liabilities", c),
        "ltDebtData": lambda c: e("long_term_debt", c),
        "nonCurrentLiabilitiesHeader": lambda c: f"{e('long_term_debt', c)}+{e('deferred_tax', c)}+{e('other_lt_liabilities', c)}",
        "totalNonCurrentLiabilities": lambda c: f"{e('long_term_debt', c)}+{e('deferred_tax', c)}+{e('other_lt_liabilities', c)}",
        "liabilitiesHeader": lambda c: e("total_liabilities", c),
        "totalLiabilities": lambda c: e("total_liabilities", c),
        "equityHeader": lambda c: e("total_equity", c),
        "totalEquity": lambda c: e("total_equity", c),
        "retainedEarnings": lambda c: e("retained_earnings", c),
        "totalLiabilitiesAndEquity": lambda c: f"{e('total_liabilities', c)}+{e('total_equity', c)}",
        # balanceCheck is handled by COMPUTED_FORMULAS in every column (A-(L+E)).
    }
    cf_ov = {
        "cfNetIncomeData": lambda c: e("net_income", c),
        "operatingActivitiesHeader": lambda c: e("operating_cash_flow", c),
        "operatingCashFlow": lambda c: e("operating_cash_flow", c),
        "capitalExpenditures": lambda c: f"-{e('capex', c)}",
        "investingActivitiesHeader": lambda c: e("investing_cash_flow", c),
        "investingCashFlow": lambda c: e("investing_cash_flow", c),
        "cfShortTermBorrowings": lambda c: e("revolver_change", c),
        "cfDividendsPaid": lambda c: f"-{e('dividends_paid', c)}",
        "financingActivitiesHeader": lambda c: e("financing_cash_flow", c),
        "financingCashFlow": lambda c: e("financing_cash_flow", c),
        "netIncreaseDecreaseCash": lambda c: e("net_cash_change", c),
        "cfEndingCashBalance": lambda c: e("cash", c),
    }
    return {"income_statement": is_ov, "balance_sheet": bs_ov, "cash_flow_statement": cf_ov}


# ============================================================
# COMPUTED SUBTOTAL/TOTAL FORMULAS
# Faithful port of frontend recalculateTotals() (calculations.js) — the app's
# single source of truth for how every header/subtotal/total is derived. Used
# to write LIVE accounting formulas (not static values) for the historical
# columns, and for the balance-check row in every column. `c(key)` returns the
# same-sheet cell reference for that line item in the current column (or "0"
# when the row isn't present).
# ============================================================
def _sum(c, keys):
    return "+".join(c(k) for k in keys)


COMPUTED_FORMULAS = {
    "income_statement": {
        "revenueHeader": lambda c: _sum(c, ["productRevenue", "serviceRevenue", "otherRevenue"]),
        "manufacturingCostsHeader": lambda c: _sum(c, ["rawMaterialsExpense", "directLabor", "manufacturingOverhead", "productionSupplies", "inventoryWriteDown"]),
        "servicesCostsHeader": lambda c: _sum(c, ["costOfServices", "professionalLabor", "projectCosts", "subcontractorCosts"]),
        "sharedCostsHeader": lambda c: c("otherCostOfRevenue"),
        "costOfRevenueDisplayHeader": lambda c: _sum(c, ["manufacturingCostsHeader", "servicesCostsHeader", "sharedCostsHeader"]),
        "grossProfit": lambda c: f"{c('revenueHeader')}-{c('costOfRevenueDisplayHeader')}",
        "sellingExpensesHeader": lambda c: _sum(c, ["sellingExpense", "advertisingAndMarketing", "distributionExpense", "otherSellingExpense"]),
        "generalAdminHeader": lambda c: _sum(c, ["generalAdminExpense", "professionalFees", "informationTechnologyExpense", "otherAdministrativeExpense"]),
        "otherOperatingExpensesHeader": lambda c: _sum(c, ["depreciationOpex", "amortizationOpex", "shareBasedCompensation", "impairmentLosses", "restructuringCharges", "otherOperatingExpense"]),
        "researchDevHeader": lambda c: c("researchDevExpense"),
        "operatingExpensesHeader": lambda c: _sum(c, ["sellingExpensesHeader", "generalAdminHeader", "researchDevHeader", "otherOperatingExpensesHeader"]),
        "operatingIncome": lambda c: f"{c('grossProfit')}-{c('operatingExpensesHeader')}",
        "nonOperatingHeader": lambda c: f"{c('financeIncome')}-{c('financeCosts')}+{c('otherNonOpIncomeExpense')}",
        "earningsBeforeTax": lambda c: f"{c('operatingIncome')}+{c('nonOperatingHeader')}",
        "incomeTaxHeader": lambda c: c("incomeTaxExpense"),
        "netIncome": lambda c: f"{c('earningsBeforeTax')}-{c('incomeTaxExpense')}",
        "netIncomeAttributableToParent": lambda c: f"{c('netIncome')}-{c('nonControllingInterest')}",
        "totalComprehensiveIncome": lambda c: f"{c('netIncome')}+{c('otherComprehensiveIncome')}",
        "ebitda": lambda c: f"{c('operatingIncome')}+{c('depreciationOpex')}+{c('amortizationOpex')}+{c('depreciationCostOfSales')}",
    },
    "balance_sheet": {
        "receivablesHeader": lambda c: f"{c('tradeAccountsReceivable')}+{c('notesReceivable')}+{c('otherReceivables')}+{c('contractAssets')}+{c('dueFromRelatedParties')}-{c('allowanceForDoubtfulAccounts')}",
        "inventoryHeader": lambda c: _sum(c, ["rawMaterials", "workInProcess", "finishedGoods", "sparePartsAndConsumables", "otherInventory"]),
        "currentAssetsHeader": lambda c: _sum(c, ["cashAndEquivalents", "restrictedCash", "shortTermInvestments", "prepaidExpenses", "vatRecoverable", "advancesToSuppliers", "receivablesHeader", "inventoryHeader", "otherCurrentAssetsData"]),
        "grossPPE": lambda c: _sum(c, ["land", "buildings", "machineryAndEquipment", "capitalWorkInProgress", "furnitureAndFixtures", "vehicles", "rightOfUseAssets", "otherPPE"]),
        "ppeHeader": lambda c: f"{c('grossPPE')}-{c('accumulatedDepreciation')}",
        "grossIntangibleAssets": lambda c: _sum(c, ["goodwill", "software", "otherIntangibleAssets"]),
        "intangibleAssetsHeader": lambda c: f"{c('grossIntangibleAssets')}-{c('accumulatedAmortization')}",
        "investmentsHeader": lambda c: _sum(c, ["equityInvestments", "debtInvestments", "investmentsInAssociates", "otherInvestments", "investmentProperty"]),
        "nonCurrentAssetsHeader": lambda c: _sum(c, ["ppeHeader", "intangibleAssetsHeader", "investmentsHeader", "deferredTaxAssets", "longTermReceivables", "otherNonCurrentAssetsData"]),
        "assetsHeader": lambda c: f"{c('currentAssetsHeader')}+{c('nonCurrentAssetsHeader')}",
        "tradePayablesHeader": lambda c: _sum(c, ["accountsPayable", "notesPayable", "otherPayables", "dueToRelatedParties"]),
        "accruedLiabilitiesHeader": lambda c: _sum(c, ["accruedExpenses", "accruedPayroll", "accruedInterest", "customerAdvances", "warrantyProvision"]),
        "currentLiabilitiesHeader": lambda c: _sum(c, ["tradePayablesHeader", "accruedLiabilitiesHeader", "stBorrowingsData", "currentPortionLTDebt", "currentLeaseLiabilities", "deferredRevenue", "contractLiabilities", "incomeTaxPayable", "vatPayable", "zakatPayable", "otherCurrentLiabilitiesData"]),
        "nonCurrentLiabilitiesHeader": lambda c: _sum(c, ["ltDebtData", "leaseLiabilities", "otherLongTermBorrowings", "deferredTaxLiabilities", "employeeEndOfServiceBenefits", "assetRetirementObligations", "otherLTLiabilitiesData"]),
        "liabilitiesHeader": lambda c: f"{c('currentLiabilitiesHeader')}+{c('nonCurrentLiabilitiesHeader')}",
        "equityHeader": lambda c: f"{c('shareCapital')}+{c('preferredStock')}+{c('additionalPaidInCapital')}-{c('treasuryStock')}+{c('retainedEarnings')}+{c('statutoryReserve')}+{c('accumulatedOCI')}+{c('otherReserves')}+{c('nonControllingInterest')}",
        "totalLiabilitiesAndEquity": lambda c: f"{c('liabilitiesHeader')}+{c('equityHeader')}",
        "balanceCheck": lambda c: f"{c('assetsHeader')}-({c('liabilitiesHeader')}+{c('equityHeader')})",
    },
    "cash_flow_statement": {
        "nonCashAdjustmentsHeader": lambda c: _sum(c, ["depreciationCostOfSales", "depreciationOpex", "amortizationOpex", "badDebtExpense", "unrealizedGainLossInvestments", "foreignExchangeGainLoss", "gainLossSaleAssets", "gainLossInvestments", "shareBasedCompensation", "deferredTax", "provisionMovements", "otherNonCashAdjustments"]),
        "workingCapitalHeader": lambda c: _sum(c, ["changeTradeAccountsReceivable", "changeContractAssetsUnbilledRevenue", "changeRelatedPartyReceivables", "changeOtherReceivables", "changeRawMaterials", "changeWorkInProcess", "changeFinishedGoods", "changeOtherInventory", "changeOtherCurrentAssets", "changeAccountsPayable", "changeOtherPayables", "changeRelatedPartyPayables", "changeAccruedExpenses", "changeCustomerAdvances", "changeDeferredRevenueContractLiab", "changeIncomeTaxPayable", "changeInterestPayable", "changeEndOfServiceBenefits", "changeOtherCurrentLiabilities", "changeOtherOperatingLiabilities"]),
        "operatingActivitiesHeader": lambda c: _sum(c, ["cfNetIncomeData", "nonCashAdjustmentsHeader", "workingCapitalHeader", "interestPaid", "interestReceived", "incomeTaxesPaid", "dividendsReceived", "otherOperatingCashFlow"]),
        "investingActivitiesHeader": lambda c: f"-{c('capitalExpenditures')}+{c('proceedsSalePPE')}-{c('purchaseInvestments')}+{c('saleInvestments')}-{c('investmentInAssociates')}-{c('purchaseIntangibleAssets')}+{c('businessAcquisitionsDisposals')}+{c('otherInvestingCashFlow')}",
        "financingActivitiesHeader": lambda c: f"{c('cfShortTermBorrowings')}+{c('cfLongTermBorrowings')}-{c('leaseLiabilityPayments')}+{c('issuanceShareCapital')}-{c('shareRepurchases')}+{c('cfAdditionalPaidInCapital')}-{c('cfDividendsPaid')}+{c('relatedPartyBorrowings')}+{c('minorityInterestTransactions')}+{c('otherFinancingCashFlow')}",
        "netIncreaseDecreaseCash": lambda c: _sum(c, ["operatingActivitiesHeader", "investingActivitiesHeader", "financingActivitiesHeader", "cfEffectOfExchangeRates"]),
        "cfEndingCashBalance": lambda c: f"{c('cfBeginningCashBalance')}+{c('netIncreaseDecreaseCash')}",
    },
}


# ============================================================
# STYLING HELPERS
# ============================================================
def _title_block(ws, title, subtitle, ncols):
    last = get_column_letter(LABEL_COL + ncols)
    ws.merge_cells(f"B1:{last}1")
    c = ws["B1"]
    c.value = title
    c.font = _font(bold=True, color=NAVY, size=14)
    c.alignment = Alignment(vertical="center")
    ws.merge_cells(f"B2:{last}2")
    s = ws["B2"]
    s.value = subtitle
    s.font = _font(color="7F7F7F", size=9)
    ws.row_dimensions[1].height = 22


def _year_header(ws, layout: StmtLayout):
    """Row 4 = ACTUAL/PROJECTED band; Row 5 = year labels."""
    hist_n = len(layout.hist_years)
    proj_n = len(layout.proj_years)
    if hist_n:
        a0 = get_column_letter(YEAR_COL0)
        a1 = get_column_letter(YEAR_COL0 + hist_n - 1)
        ws.merge_cells(f"{a0}4:{a1}4")
        cc = ws[f"{a0}4"]
        cc.value = "ACTUAL"
        cc.font = _font(bold=True, color="7F7F7F", size=9)
        cc.alignment = Alignment(horizontal="center")
    if proj_n:
        p0 = get_column_letter(YEAR_COL0 + hist_n)
        p1 = get_column_letter(YEAR_COL0 + hist_n + proj_n - 1)
        ws.merge_cells(f"{p0}4:{p1}4")
        cc = ws[f"{p0}4"]
        cc.value = "PROJECTED"
        cc.font = _font(bold=True, color=NAVY_MED, size=9)
        cc.alignment = Alignment(horizontal="center")
        cc.fill = PatternFill("solid", fgColor=ACCENT_PROJ)

    hdr = ws.cell(row=5, column=LABEL_COL, value="SAR '000")
    hdr.font = _font(bold=True, color=WHITE, size=9)
    hdr.fill = PatternFill("solid", fgColor=NAVY)
    hdr.alignment = Alignment(horizontal="left")
    for i, y in enumerate(layout.years):
        col = YEAR_COL0 + i
        is_proj = y in layout.proj_years
        suffix = "P" if is_proj else "A"
        cc = ws.cell(row=5, column=col, value=f"{_parse_year(y)}{suffix}")
        cc.font = _font(bold=True, color=WHITE, size=10)
        cc.fill = PatternFill("solid", fgColor=NAVY if not is_proj else NAVY_MED)
        cc.alignment = Alignment(horizontal="center")


def _apply_col_widths(ws, ncols, label_width=34):
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions[get_column_letter(LABEL_COL)].width = label_width
    for i in range(ncols):
        ws.column_dimensions[get_column_letter(YEAR_COL0 + i)].width = 12


# ============================================================
# STATEMENT SHEET
# ============================================================
def _build_statement(ws, layout: StmtLayout, subtitle, overrides_for_stmt, engine_row,
                     mult_name="rev_mult", stmt_type="income_statement"):
    ncols = len(layout.years)
    _title_block(ws, layout.title, subtitle, ncols)
    _year_header(ws, layout)
    _apply_col_widths(ws, ncols)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = f"{get_column_letter(YEAR_COL0)}6"

    E = _q(MODEL_SHEET)
    computed = COMPUTED_FORMULAS.get(stmt_type, {})

    def cref(col):
        """Same-sheet cell ref for a key in `col` (or '0' if the row is absent)."""
        return lambda k: (f"{col}{layout.key_row[k]}" if k in layout.key_row else "0")

    for i, r in enumerate(layout.template_rows):
        row = DATA_ROW0 + i
        key = r["key"]
        level = r.get("level", 3)
        is_header = r.get("is_header", False)
        is_sub = r.get("is_subtotal", False)
        indent = max(0, level - 1)
        is_computed = key in computed          # header/subtotal/total with a formula
        # balanceCheck is a live A-(L+E) formula in EVERY column
        computed_all_cols = key == "balanceCheck"

        # ── label
        lc = ws.cell(row=row, column=LABEL_COL, value=r["label"])
        lc.alignment = Alignment(indent=indent)
        if is_header:
            lc.font = _font(bold=True, color=NAVY, size=10)
        elif is_sub:
            lc.font = _font(bold=True, color=BLACK, size=10)
        else:
            lc.font = _font(color="333333", size=10)

        # section-header shading across the row
        if is_header:
            for cidx in range(LABEL_COL, YEAR_COL0 + ncols):
                ws.cell(row=row, column=cidx).fill = PatternFill("solid", fgColor=GREY_HEADER)

        vals = layout.stored_lookup.get(key, {})

        # ── historical columns ──
        # Headers/subtotals/totals = LIVE accounting formulas (port of the app's
        # recalculateTotals); detail lines = stored actuals (blue inputs).
        for y in layout.hist_years:
            col = layout.year_col[y]
            cidx = ws[f"{col}{row}"].column
            cell = ws.cell(row=row, column=cidx)
            cell.number_format = FMT_NUM
            if is_computed:
                cell.value = "=" + computed[key](cref(col))
                cell.font = _font(bold=is_sub or is_header, color=BLACK)
            else:
                v = vals.get(y)
                if v is not None and v != 0:
                    cell.value = float(v)
                    cell.font = _font(bold=is_sub, color=BLUE_INPUT if not is_header else NAVY)

        # ── projected columns: live driver-based model ──
        for k, y in enumerate(layout.proj_years):
            col = layout.year_col[y]
            eng_col = get_column_letter(ENGINE_F0 + k)   # engine forecast column
            cidx = ws[f"{col}{row}"].column
            cell = ws.cell(row=row, column=cidx)
            cell.number_format = FMT_NUM
            if computed_all_cols:
                cell.value = "=" + computed[key](cref(col))
                cell.font = _font(bold=is_sub, color=BLACK)
            elif key in overrides_for_stmt:
                cell.value = "=" + overrides_for_stmt[key](eng_col)
                cell.font = _font(bold=is_sub, color=GREEN_LINK)
            else:
                # proportional scaling: base_cell * multiplier (matches
                # run_forecast: IS/CFS scale by revenue, BS by total assets)
                base_cell = f"{layout.base_col}{row}" if layout.base_col else "0"
                mult_ref = f"{E}!{eng_col}{engine_row[mult_name]}"
                cell.value = f"={base_cell}*{mult_ref}"
                cell.font = _font(bold=is_sub, color=BLACK)

        # subtotal top rule
        if is_sub:
            for cidx in range(YEAR_COL0, YEAR_COL0 + ncols):
                cc = ws.cell(row=row, column=cidx)
                cc.border = Border(top=THIN)


# First engine forecast column on the merged Assumptions sheet: the engine
# block uses the sheet's own layout (labels=B, base=C, forecast=D..).
ENGINE_F0 = YEAR_COL0 + 1


# ============================================================
# ENGINE BLOCK  (calculation rows on the Assumptions sheet, below the
# driver panel — a live-formula port of forecasting_engine, grouped and
# collapsed under a "do not edit" banner)
# ============================================================
def _build_engine(ws, engine_row, base, assum_row, n_forecast, balance_mode,
                  banner_row, base_year, proj_years):
    """
    Writes the engine variables onto the Assumptions sheet using its own
    column layout: labels=B, column C = base-year anchors, columns D.. =
    forecast years, each referencing the prior column and the driver cells
    above — a line-for-line port of generate_forecast().
    """
    faithful = balance_mode == "faithful"
    BASE_L = get_column_letter(YEAR_COL0)          # base column letter ("C")
    last_col = YEAR_COL0 + n_forecast

    # ── banner + block header ──
    bc = ws.cell(row=banner_row, column=LABEL_COL,
                 value="FORECAST ENGINE — CALCULATED (do not edit; driven by the drivers above)")
    bc.font = _font(bold=True, color=WHITE, size=10)
    for cidx in range(LABEL_COL, last_col + 1):
        ws.cell(row=banner_row, column=cidx).fill = PatternFill("solid", fgColor=NAVY)
    hr = banner_row + 1
    ws.cell(row=hr, column=LABEL_COL, value="Engine Variable").font = \
        _font(bold=True, color=NAVY_MED, size=9)
    ws.cell(row=hr, column=YEAR_COL0, value=f"Base {base_year}A").font = \
        _font(bold=True, color=NAVY_MED, size=9)
    for k, y in enumerate(proj_years):
        cc = ws.cell(row=hr, column=YEAR_COL0 + 1 + k, value=f"{_parse_year(y)}P")
        cc.font = _font(bold=True, color=NAVY_MED, size=9)
        cc.alignment = Alignment(horizontal="center")

    # Label column + number formats (grey, small — visually "calculated")
    for name, label, fmt in ENGINE_VARS:
        r = engine_row[name]
        lc = ws.cell(row=r, column=LABEL_COL, value=label)
        lc.font = _font(size=9, color="555555")
        lc.alignment = Alignment(indent=1)
        for cidx in range(YEAR_COL0, last_col + 1):
            cc = ws.cell(row=r, column=cidx)
            cc.number_format = fmt
            cc.font = _font(size=9, color="555555")

    def R(name):
        return engine_row[name]

    # helper to write base (col C) and forecast (cols D..) formulas
    def put(name, base_formula, fc_formula):
        r = R(name)
        ws[f"{BASE_L}{r}"] = base_formula
        for k in range(n_forecast):
            col = get_column_letter(YEAR_COL0 + 1 + k)
            prev = get_column_letter(YEAR_COL0 + k)   # column to the left (base for k=0)
            ws[f"{col}{r}"] = fc_formula(col, prev, col)

    # ── Base anchors ──
    # Computed in Python exactly as forecasting_engine.extract_base_data (which
    # applies the compat/sum key resolution — including legacy keys that aren't
    # in the current template, e.g. a raw "costOfRevenue" row). Written as
    # literals in the Base column so the projected model reproduces the app's
    # engine for EVERY project. Projections stay fully live via the editable
    # Assumptions drivers; only the base-year starting point is fixed.
    base_revenue = base.revenue
    base_cost = base.cost_of_revenue
    base_opinc = base.operating_income
    base_dep = base.depreciation
    base_inv = base.inventory
    base_ar = base.accounts_receivable
    base_ap = base.accounts_payable
    base_cash = base.cash
    base_ta = base.total_assets
    base_tl = base.total_liabilities
    base_te = base.total_equity
    base_ca = base.current_assets
    base_cl = base.current_liabilities
    base_ltd = base.long_term_debt
    base_re = base.retained_earnings
    base_ppe = base.ppe
    base_accdep = base.accumulated_depreciation

    # Driver cell on this same sheet — the engine block shares the driver
    # panel's column layout, so the reference is just column letter + row.
    def drv(name, col):
        return f"{col}{assum_row[name]}"

    # ── Cost-structure ratios (per-year, read from the Assumptions sheet;
    #    col B keeps the base-year derivation for reference) ──
    put("cogs_pct",
        f"=IF({base_revenue}=0,0.6,{base_cost}/{base_revenue})",
        lambda c, p, a: f"={drv('cogs_pct', a)}/100")
    # sga_pct base: engine defaults to 0.25 when base revenue is 0 (guard against #DIV/0)
    put("sga_pct",
        f"=IF({base_revenue}=0,0.25,"
        f"MAX(0.05,({base_opinc}/{base_revenue})-(1-C{R('cogs_pct')}-{drv('dep_rate','C')}/100)+0.08))",
        lambda c, p, a: f"={drv('sga_pct', a)}/100")
    put("rd_pct", "=0.08",
        lambda c, p, a: f"={drv('rd_pct', a)}/100")

    # ── Revenue & IS ──
    put("revenue",
        f"={base_revenue}",
        lambda c, p, a: f"={p}{R('revenue')}*(1+{drv('growth', a)}/100)")
    put("rev_mult",
        "=1",
        lambda c, p, a: f"=IF($C${R('revenue')}=0,1,{c}{R('revenue')}/$C${R('revenue')})")
    put("ast_mult",
        "=1",
        lambda c, p, a: f"=IF($C${R('total_assets')}=0,1,{c}{R('total_assets')}/$C${R('total_assets')})")
    put("cost_of_revenue", f"={base_cost}",
        lambda c, p, a: f"={c}{R('revenue')}*{c}{R('cogs_pct')}")
    put("gross_profit", f"={base_revenue}-{base_cost}",
        lambda c, p, a: f"={c}{R('revenue')}-{c}{R('cost_of_revenue')}")
    put("depreciation", f"={base_dep}",
        lambda c, p, a: f"={c}{R('revenue')}*{drv('dep_rate', a)}/100")
    put("sga_expenses", "=0",
        lambda c, p, a: f"={c}{R('revenue')}*{c}{R('sga_pct')}")
    put("rd_expenses", "=0",
        lambda c, p, a: f"={c}{R('revenue')}*{c}{R('rd_pct')}")
    put("total_op_expenses", "=0",
        lambda c, p, a: f"={c}{R('sga_expenses')}+{c}{R('rd_expenses')}+{c}{R('depreciation')}")
    put("operating_income", f"={base_opinc}",
        lambda c, p, a: f"={c}{R('gross_profit')}-{c}{R('total_op_expenses')}")
    put("ebitda", f"={base_opinc}+{base_dep}",
        lambda c, p, a: f"={c}{R('operating_income')}+{c}{R('depreciation')}")
    put("interest_income", "=0",
        lambda c, p, a: f"={p}{R('cash')}*{drv('int_income', a)}/100")
    put("interest_expense", "=0",
        lambda c, p, a: f"=({c}{R('long_term_debt')}+{p}{R('revolver')})*{drv('interest', a)}/100")
    put("income_before_tax", "=0",
        lambda c, p, a: f"={c}{R('operating_income')}+{c}{R('interest_income')}-{c}{R('interest_expense')}")
    put("tax_expense", "=0",
        lambda c, p, a: f"={c}{R('income_before_tax')}*{drv('tax', a)}/100")
    put("net_income", "=0",
        lambda c, p, a: f"={c}{R('income_before_tax')}-{c}{R('tax_expense')}")

    # ── Working capital (ratio-driven) ──
    put("accounts_receivable", f"={base_ar}",
        lambda c, p, a: f"={c}{R('revenue')}*{drv('dso', a)}/365")
    put("inventory", f"={base_inv}",
        lambda c, p, a: f"={c}{R('cost_of_revenue')}*{drv('dio', a)}/365")
    put("accounts_payable", f"={base_ap}",
        lambda c, p, a: f"={c}{R('cost_of_revenue')}*{drv('dpo', a)}/365")
    put("capex", "=0",
        lambda c, p, a: f"={c}{R('revenue')}*{drv('capex', a)}/100")
    put("cumulative_ppe", f"={base_ppe}",
        lambda c, p, a: f"={p}{R('cumulative_ppe')}+{c}{R('capex')}")
    put("accumulated_depreciation", f"={base_accdep}",
        lambda c, p, a: f"={p}{R('accumulated_depreciation')}+{c}{R('depreciation')}")
    put("net_ppe", f"={base_ppe}-{base_accdep}",
        lambda c, p, a: f"={c}{R('cumulative_ppe')}-{c}{R('accumulated_depreciation')}")
    put("wc_adjustment", "=0",
        lambda c, p, a: (f"=-(({c}{R('accounts_receivable')}-{p}{R('accounts_receivable')})"
                         f"+({c}{R('inventory')}-{p}{R('inventory')})"
                         f"-({c}{R('accounts_payable')}-{p}{R('accounts_payable')}))"))
    put("common_stock", f"=IF(({base_te}-{base_re})=0,100000000,{base_te}-{base_re})",
        lambda c, p, a: f"=$C${R('common_stock')}")
    put("dividends_paid", "=0",
        lambda c, p, a: f"=MAX(0,{c}{R('net_income')}*{drv('dividend', a)}/100)")
    put("retained_earnings", f"={base_re}",
        lambda c, p, a: f"={p}{R('retained_earnings')}+{c}{R('net_income')}-{c}{R('dividends_paid')}")
    put("total_equity", f"={base_te}",
        lambda c, p, a: f"={c}{R('common_stock')}+{c}{R('retained_earnings')}")
    put("long_term_debt", f"={base_ltd}",
        lambda c, p, a: f"=$C${R('long_term_debt')}")

    if not faithful:
        # ── BALANCED MODE ──
        put("other_current_assets", "=0",
            lambda c, p, a: f"={c}{R('revenue')}*{drv('other_ca', a)}/100")
        put("acquisitions", "=0",
            lambda c, p, a: f"={c}{R('revenue')}*{drv('acq', a)}/100")
        put("goodwill", "=0",
            lambda c, p, a: f"={base_ta}*{drv('goodwill', a)}/100")
        put("other_intangibles", "=0",
            lambda c, p, a: f"={base_ta}*{drv('intang', a)}/100")
        put("non_cash_assets", "=0",
            lambda c, p, a: (f"=({c}{R('accounts_receivable')}+{c}{R('inventory')}+{c}{R('other_current_assets')})"
                             f"+{c}{R('net_ppe')}+{c}{R('goodwill')}+{c}{R('other_intangibles')}"))
        put("other_current_liabilities", "=0",
            lambda c, p, a: f"={c}{R('revenue')}*{drv('other_cl', a)}/100")
        put("deferred_tax", "=0",
            lambda c, p, a: f"={c}{R('non_cash_assets')}*{drv('dt', a)}/100")
        put("other_lt_liabilities", "=0",
            lambda c, p, a: f"={c}{R('non_cash_assets')}*{drv('oltl', a)}/100")
        put("pre_revolver_liabilities", "=0",
            lambda c, p, a: (f"={c}{R('accounts_payable')}+{c}{R('other_current_liabilities')}"
                             f"+{c}{R('long_term_debt')}+{c}{R('deferred_tax')}+{c}{R('other_lt_liabilities')}"))
        put("cash_surplus", "=0",
            lambda c, p, a: f"=({c}{R('pre_revolver_liabilities')}+{c}{R('total_equity')})-{c}{R('non_cash_assets')}")
        put("cash", f"={base_cash}",
            lambda c, p, a: f"=IF({c}{R('cash_surplus')}>=0,{c}{R('cash_surplus')},0)")
        put("revolver", "=0",
            lambda c, p, a: f"=IF({c}{R('cash_surplus')}>=0,0,-{c}{R('cash_surplus')})")
        put("revolver_change", "=0",
            lambda c, p, a: f"={c}{R('revolver')}-{p}{R('revolver')}")
        put("total_current_liabilities", "=0",
            lambda c, p, a: f"={c}{R('accounts_payable')}+{c}{R('other_current_liabilities')}+{c}{R('revolver')}")
        put("total_liabilities", "=0",
            lambda c, p, a: f"={c}{R('pre_revolver_liabilities')}+{c}{R('revolver')}")
        put("total_current_assets", "=0",
            lambda c, p, a: (f"={c}{R('cash')}+{c}{R('accounts_receivable')}"
                             f"+{c}{R('inventory')}+{c}{R('other_current_assets')}"))
        put("total_assets", "=0",
            lambda c, p, a: (f"={c}{R('total_current_assets')}+{c}{R('net_ppe')}"
                             f"+{c}{R('goodwill')}+{c}{R('other_intangibles')}"))
        put("stock_based_comp", "=0",
            lambda c, p, a: f"={c}{R('revenue')}*{drv('sbc', a)}/100")
        put("deferred_tax_change", "=0",
            lambda c, p, a: f"={c}{R('non_cash_assets')}*{drv('dtc', a)}/100")
        put("operating_cash_flow", "=0",
            lambda c, p, a: (f"={c}{R('net_income')}+{c}{R('depreciation')}+{c}{R('stock_based_comp')}"
                             f"+{c}{R('deferred_tax_change')}+{c}{R('wc_adjustment')}"))
        put("investing_cash_flow", "=0",
            lambda c, p, a: f"=-({c}{R('capex')}+{c}{R('acquisitions')})")
        put("debt_issuance", "=0",
            lambda c, p, a: f"={c}{R('revenue')}*{drv('debt_iss', a)}/100")
        put("financing_cash_flow", "=0",
            lambda c, p, a: f"=-{c}{R('dividends_paid')}+{c}{R('debt_issuance')}+{c}{R('revolver_change')}")
        put("net_cash_change", "=0",
            lambda c, p, a: f"={c}{R('cash')}-{p}{R('cash')}")
    else:
        # ── FAITHFUL MODE ──
        base_other_ca = f"({base_ca}-{base_cash}-({base_ar})-({base_inv}))"
        base_other_nca = f"({base_ta}-{base_ca}-({base_ppe}-{base_accdep}))"
        base_other_cl = f"({base_cl}-{base_ap})"
        base_other_ltl = f"({base_tl}-{base_cl}-{base_ltd})"
        put("other_current_assets", "=0", lambda c, p, a: f"={base_other_ca}")
        put("acquisitions", "=0", lambda c, p, a: "=0")
        put("goodwill", "=0", lambda c, p, a: f"={base_other_nca}")
        put("other_intangibles", "=0", lambda c, p, a: "=0")
        put("other_current_liabilities", "=0", lambda c, p, a: f"={base_other_cl}")
        put("deferred_tax", "=0", lambda c, p, a: "=0")
        put("other_lt_liabilities", "=0", lambda c, p, a: f"={base_other_ltl}")
        put("non_cash_assets", "=0",
            lambda c, p, a: (f"=({c}{R('accounts_receivable')}+{c}{R('inventory')}+{c}{R('other_current_assets')})"
                             f"+{c}{R('net_ppe')}+{c}{R('goodwill')}"))
        put("stock_based_comp", "=0", lambda c, p, a: "=0")
        put("deferred_tax_change", "=0", lambda c, p, a: "=0")
        put("operating_cash_flow", "=0",
            lambda c, p, a: f"={c}{R('net_income')}+{c}{R('depreciation')}+{c}{R('wc_adjustment')}")
        put("investing_cash_flow", "=0", lambda c, p, a: f"=-{c}{R('capex')}")
        put("debt_issuance", "=0", lambda c, p, a: "=0")
        put("financing_cash_flow", "=0", lambda c, p, a: f"=-{c}{R('dividends_paid')}")
        put("net_cash_change", "=0",
            lambda c, p, a: f"={c}{R('operating_cash_flow')}+{c}{R('investing_cash_flow')}+{c}{R('financing_cash_flow')}")
        put("cash", f"={base_cash}",
            lambda c, p, a: f"={p}{R('cash')}+{c}{R('net_cash_change')}")
        put("revolver", "=0", lambda c, p, a: "=0")
        put("revolver_change", "=0", lambda c, p, a: "=0")
        put("pre_revolver_liabilities", "=0", lambda c, p, a: "=0")
        put("cash_surplus", "=0", lambda c, p, a: "=0")
        put("total_current_assets", "=0",
            lambda c, p, a: (f"={c}{R('cash')}+{c}{R('accounts_receivable')}"
                             f"+{c}{R('inventory')}+{c}{R('other_current_assets')}"))
        put("total_assets", "=0",
            lambda c, p, a: f"={c}{R('total_current_assets')}+{c}{R('net_ppe')}+{c}{R('goodwill')}")
        put("total_current_liabilities", "=0",
            lambda c, p, a: f"={c}{R('accounts_payable')}+{c}{R('other_current_liabilities')}")
        put("total_liabilities", "=0",
            lambda c, p, a: f"={c}{R('total_current_liabilities')}+{c}{R('long_term_debt')}+{c}{R('other_lt_liabilities')}")

    put("free_cash_flow", "=0",
        lambda c, p, a: f"={c}{R('operating_cash_flow')}-{c}{R('capex')}")

    # Group the whole block (header row + variables) one outline level deep,
    # collapsed by default — the banner row stays visible with a [+] control.
    ws.sheet_properties.outlinePr.summaryBelow = False
    last_row = max(engine_row.values())
    for rr in range(banner_row + 1, last_row + 1):
        ws.row_dimensions[rr].outline_level = 1
        ws.row_dimensions[rr].hidden = True
    ws.row_dimensions[banner_row].collapsed = True


# ============================================================
# ASSUMPTIONS SHEET  (editable driver panel — blue inputs)
# ============================================================
# Every driver the projected model uses is exposed here — including the
# cost-structure and balance-sheet-build ratios the engine previously kept
# internal — one editable blue cell per forecast year. The Base column shows
# the base-year actual as a live (italic) formula off the statement sheets;
# where a driver has no historical counterpart the model default is shown.
# The hidden Forecast Engine sheet reads these cells, so this is the single
# editing surface for the whole model.


def _build_assumptions(ws, company, base_year, proj_years, inputs, base,
                       is_l, bs_l, cf_l, balance_mode,
                       mode_label="", dcf_assumptions=None):
    n = len(proj_years)
    ncols = 1 + n  # base + forecast
    _apply_col_widths(ws, ncols, label_width=42)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = f"{get_column_letter(YEAR_COL0)}6"

    _title_block(ws, f"{company} — ASSUMPTIONS & DRIVERS", ncols=ncols,
                 subtitle="Projection inputs (blue) drive the entire model. Base year = "
                          f"FY{base_year}. Edit the blue cells to re-forecast. "
                          "Italic = base-year actual (reference only).")
    if mode_label:
        mc = ws.cell(row=3, column=LABEL_COL,
                     value=f"Projection basis:  {mode_label}")
        mc.font = _font(bold=True, italic=True, color=NAVY_MED, size=10)
    # year header
    ws.cell(row=5, column=LABEL_COL, value="Driver").font = _font(bold=True, color=WHITE)
    ws.cell(row=5, column=LABEL_COL).fill = PatternFill("solid", fgColor=NAVY)
    ws.cell(row=5, column=YEAR_COL0, value=f"Base {base_year}A").font = _font(bold=True, color=WHITE)
    ws.cell(row=5, column=YEAR_COL0).fill = PatternFill("solid", fgColor=NAVY)
    for k, y in enumerate(proj_years):
        cc = ws.cell(row=5, column=YEAR_COL0 + 1 + k, value=f"{_parse_year(y)}P")
        cc.font = _font(bold=True, color=WHITE)
        cc.fill = PatternFill("solid", fgColor=NAVY_MED)
        cc.alignment = Alignment(horizontal="center")

    # per-year growth values (carry last forward)
    grates = inputs.get("revenue_growth_rates")

    def year_val(name, k):
        if name == "growth":
            if grates:
                return float(grates[min(k, len(grates) - 1)])
            return float(inputs.get("revenue_growth_rate", 10.0))
        keymap = {
            "tax": "tax_rate", "capex": "capex_as_pct_of_revenue",
            "dividend": "dividend_payout_ratio", "interest": "interest_rate_on_debt",
            "dso": "dso", "dio": "dio", "dpo": "dpo", "dep_rate": "depreciation_rate",
        }
        return float(inputs.get(keymap[name], 0.0))

    # ── live base-column refs (base-year actuals from the statement sheets) ──
    def isc(key): return is_l.cell(key, is_l.base_col)
    def bsc(key): return bs_l.cell(key, bs_l.base_col)
    def cfc(key): return cf_l.cell(key, cf_l.base_col)

    rev = isc("revenueHeader")
    cogs = f"ABS({isc('costOfRevenueDisplayHeader')})"
    opinc = isc("operatingIncome")
    debt = f"({bsc('ltDebtData')}+{bsc('stBorrowingsData')}+{bsc('currentPortionLTDebt')})"
    # template uses "earningsBeforeTax"; accept the legacy key too
    ibt = isc("earningsBeforeTax") if is_l.has("earningsBeforeTax") else isc("incomeBeforeTax")

    # Cost-structure defaults: the exact (unrounded) values forecasting_engine
    # derives from the base year, so the default workbook reproduces the app.
    rev0 = base.revenue
    cogs_def = (base.cost_of_revenue / rev0 * 100) if rev0 else 60.0
    dep_in = float(inputs.get("depreciation_rate", 8.0))
    sga_def = (max(0.05, (base.operating_income / rev0)
                   - (1 - cogs_def / 100 - dep_in / 100) + 0.08) * 100) if rev0 else 25.0

    def const(v):
        return lambda k: v

    base_col_letter = get_column_letter(YEAR_COL0)

    def sga_base(rows):
        # Mirrors the engine's SG&A derivation, against this sheet's own
        # depreciation base cell.
        return (f'=IF({rev}=0,"",MAX(5,({opinc}/{rev}-(1-{cogs}/{rev}'
                f'-{base_col_letter}{rows["dep_rate"]}/100)+0.08)*100))')

    growth_base = None
    if len(is_l.hist_years) >= 2:
        prev = is_l.cell("revenueHeader", is_l.year_col[is_l.hist_years[-2]])
        growth_base = f'=IF({prev}=0,"",({rev}/{prev}-1)*100)'

    # (name, label, fmt, base cell, per-year default) — base cell is a live
    # formula string, a literal default (no historical counterpart), a
    # callable(assum_row) for forward row refs, or None to leave blank.
    sections = [
        ("REVENUE DRIVERS", [
            ("growth", "Revenue Growth Rate (YoY %)", FMT_NUM1, growth_base,
             lambda k: year_val("growth", k)),
        ]),
        ("COST STRUCTURE", [
            ("cogs_pct", "Cost of Revenue / Revenue (%)", FMT_NUM1,
             f'=IF({rev}=0,"",{cogs}/{rev}*100)', const(cogs_def)),
            ("sga_pct", "SG&A Expense / Revenue (%)", FMT_NUM1, sga_base, const(sga_def)),
            ("rd_pct", "R&D Expense / Revenue (%)", FMT_NUM1, 8.0, const(8.0)),
        ]),
        ("CAPEX & DEPRECIATION", [
            ("capex", "CapEx (% of Revenue)", FMT_NUM1,
             f'=IF({rev}=0,"",ABS({cfc("capitalExpenditures")})/{rev}*100)',
             lambda k: year_val("capex", k)),
            ("dep_rate", "Depreciation (% of Revenue)", FMT_NUM1,
             f'=IF({rev}=0,"",ABS({isc("depreciationOpex")})/{rev}*100)',
             lambda k: year_val("dep_rate", k)),
        ]),
        ("WORKING CAPITAL DRIVERS", [
            ("dso", "Days Sales Outstanding (DSO)", FMT_DAYS,
             f'=IF({rev}=0,"",{bsc("receivablesHeader")}/{rev}*365)',
             lambda k: year_val("dso", k)),
            ("dio", "Days Inventory Outstanding (DIO)", FMT_DAYS,
             f'=IF({cogs}=0,"",{bsc("inventoryHeader")}/{cogs}*365)',
             lambda k: year_val("dio", k)),
            ("dpo", "Days Payable Outstanding (DPO)", FMT_DAYS,
             f'=IF({cogs}=0,"",{bsc("tradePayablesHeader")}/{cogs}*365)',
             lambda k: year_val("dpo", k)),
        ]),
        ("FINANCING & TAX", [
            ("interest", "Interest Rate on Debt (%)", FMT_NUM1,
             f'=IF({debt}=0,"",ABS({isc("financeCosts")})/{debt}*100)',
             lambda k: year_val("interest", k)),
            ("int_income", "Interest Income (% of Prior-Year Cash)", FMT_NUM1,
             2.0, const(2.0)),
            ("tax", "Tax Rate (%)", FMT_NUM1,
             f'=IF({ibt}=0,"",ABS({isc("incomeTaxExpense")})/{ibt}*100)',
             lambda k: year_val("tax", k)),
            ("dividend", "Dividend Payout Ratio (%)", FMT_NUM1,
             f'=IF({isc("netIncome")}=0,"",'
             f'ABS({cfc("cfDividendsPaid")})/{isc("netIncome")}*100)',
             lambda k: year_val("dividend", k)),
        ]),
    ]

    if balance_mode != "faithful":
        # Balanced mode builds the unmodeled BS/CF lines from ratios that used
        # to be hardcoded on the hidden engine sheet — now editable here.
        ca, cl = bsc("currentAssetsHeader"), bsc("currentLiabilitiesHeader")
        cash = bsc("cashAndEquivalents")
        ar, inv, ap = bsc("receivablesHeader"), bsc("inventoryHeader"), bsc("tradePayablesHeader")
        sections.append(("BALANCE SHEET & CASH FLOW BUILD (Balanced mode)", [
            ("other_ca", "Other Current Assets (% of Revenue)", FMT_NUM1,
             f'=IF({rev}=0,"",({ca}-{cash}-{ar}-{inv})/{rev}*100)', const(5.0)),
            ("other_cl", "Other Current Liabilities (% of Revenue)", FMT_NUM1,
             f'=IF({rev}=0,"",({cl}-{ap})/{rev}*100)', const(3.0)),
            ("acq", "Acquisitions (% of Revenue)", FMT_NUM1, 1.0, const(1.0)),
            ("goodwill", "Goodwill / Other NCA (% of Base Total Assets)", FMT_NUM1,
             10.0, const(10.0)),
            ("intang", "Other Intangibles (% of Base Total Assets)", FMT_NUM1,
             5.0, const(5.0)),
            ("dt", "Deferred Tax (% of Non-cash Assets)", FMT_NUM1, 2.0, const(2.0)),
            ("oltl", "Other LT Liabilities (% of Non-cash Assets)", FMT_NUM1,
             3.0, const(3.0)),
            ("sbc", "Stock-Based Comp (% of Revenue)", FMT_NUM1, 2.0, const(2.0)),
            ("dtc", "Deferred Tax Change (% of Non-cash Assets)", FMT_NUM1,
             0.1, const(0.1)),
            ("debt_iss", "Debt Issuance (% of Revenue)", FMT_NUM1, None,
             lambda k: 5.0 if k == 0 else 0.0),
        ]))

    # first pass: assign a row to every driver so base formulas can reference
    # other driver rows regardless of order
    assum_row = {}
    row_plan = []
    r = 6
    for title, drivers in sections:
        row_plan.append(("section", title, r))
        r += 1
        for d in drivers:
            assum_row[d[0]] = r
            row_plan.append(("driver", d, r))
            r += 1

    for kind, payload, rr in row_plan:
        if kind == "section":
            cc = ws.cell(row=rr, column=LABEL_COL, value=payload)
            cc.font = _font(bold=True, color=NAVY, size=10)
            for cidx in range(LABEL_COL, YEAR_COL0 + ncols):
                ws.cell(row=rr, column=cidx).fill = PatternFill("solid", fgColor=GREY_HEADER)
            continue
        name, label, fmt, base_val, val_fn = payload
        ws.cell(row=rr, column=LABEL_COL, value=label).font = _font(size=10, color="333333")
        ws.cell(row=rr, column=LABEL_COL).alignment = Alignment(indent=1)
        # base column (italic reference — live actual or model default)
        bc = ws.cell(row=rr, column=YEAR_COL0)
        bc.number_format = fmt
        if callable(base_val):
            base_val = base_val(assum_row)
        if base_val is not None:
            bc.value = base_val
            bc.font = _font(size=10, italic=True, color="555555")
        # forecast columns (editable blue inputs)
        for k in range(n):
            cc = ws.cell(row=rr, column=YEAR_COL0 + 1 + k, value=val_fn(k))
            cc.number_format = fmt
            cc.font = _font(color=BLUE_INPUT)

    # ── Valuation (DCF) drivers — scalar, single value, not per-year ──
    # Merged onto this sheet so the forecast and the DCF share one uniform
    # assumptions surface instead of the DCF sheet holding its own copy.
    cc = ws.cell(row=r, column=LABEL_COL, value="VALUATION (DCF) DRIVERS")
    cc.font = _font(bold=True, color=NAVY, size=10)
    for cidx in range(LABEL_COL, YEAR_COL0 + ncols):
        ws.cell(row=r, column=cidx).fill = PatternFill("solid", fgColor=GREY_HEADER)
    r += 1

    da = dcf_assumptions or {}
    wacc = float(da.get("wacc", 10) or 10)
    tgr = float(da.get("terminal_growth_rate", 2.5) or 2.5)
    shares = da.get("shares_outstanding", "")
    try:
        shares = float(shares)
    except (TypeError, ValueError):
        shares = 0.0

    dcf_row = {}
    for name, label, fmt, default in [
        ("wacc", "WACC (%)", FMT_NUM1, wacc),
        ("tgr", "Terminal Growth Rate (%)", FMT_NUM1, tgr),
        ("method", "Valuation Method", "General", "Perpetuity"),
        ("exit_mult", "Exit Multiple (x)", FMT_RATIO, 12.0),
        ("shares", "Shares Outstanding", FMT_NUM, shares),
        ("fcst_years", "Forecast Years", "0", n),
    ]:
        dcf_row[name] = r
        ws.cell(row=r, column=LABEL_COL, value=label).font = _font(size=10, color="333333")
        ws.cell(row=r, column=LABEL_COL).alignment = Alignment(indent=1)
        cc = ws.cell(row=r, column=YEAR_COL0, value=default)
        cc.number_format = fmt
        cc.font = _font(color=BLUE_INPUT, bold=True)
        r += 1

    dv = DataValidation(type="list", formula1='"Perpetuity,Exit Multiple"', allow_blank=False)
    ws.add_data_validation(dv)
    dv.add(ws.cell(row=dcf_row["method"], column=YEAR_COL0))

    note = ws.cell(row=r + 1, column=LABEL_COL,
                   value="Blue = editable input. Change any driver to update every "
                         "statement, ratio, and the DCF.")
    note.font = _font(italic=True, size=9, color="7F7F7F")
    # r+2 stays blank as a spacer; the engine block's banner goes at r+3.
    return assum_row, dcf_row, r + 3


# ============================================================
# RATIOS SHEET  (live cross-sheet formulas, all years)
# ============================================================
def _build_ratios(ws, company, is_l, bs_l, cf_l):
    years = is_l.years
    ncols = len(years)
    _apply_col_widths(ws, ncols, label_width=34)
    ws.sheet_view.showGridLines = False
    _title_block(ws, f"{company} — FINANCIAL RATIOS", ncols=ncols,
                 subtitle="Live formulas linked to the statement sheets · Actual & Projected")
    # year header (reuse statement-style header on is_l years)
    _ratio_year_header(ws, years, is_l)

    def C(layout, key, col):
        return layout.cell(key, col)

    def col_of(k):
        return get_column_letter(YEAR_COL0 + k)

    # per-year cell shorthands
    def refs(k, layout_years):
        col = is_l.year_col[layout_years[k]]
        return col

    rows = []  # (category, label, fmt, fn(k)->formula body without '=')

    def add(cat, label, fmt, fn):
        rows.append((cat, label, fmt, fn))

    def isc(key, col): return is_l.cell(key, col)
    def bsc(key, col): return bs_l.cell(key, col)
    def cfc(key, col): return cf_l.cell(key, col)

    def ycol(k): return is_l.year_col[years[k]]

    def taxrate(col):
        ebt = f"{isc('earningsBeforeTax', col)}"
        tax = f"ABS({isc('incomeTaxExpense', col)})"
        return f"IF(AND({ebt}>0,{tax}/{ebt}>=0,{tax}/{ebt}<=1),{tax}/{ebt},0.25)"

    def avg(layout, key, k):
        cur = layout.cell(key, ycol(k))
        if k == 0:
            return cur
        prev = layout.cell(key, ycol(k - 1))
        return f"(({cur})+({prev}))/2"

    # Build metric definitions as functions of k (year index)
    def rev(k): return isc("revenueHeader", ycol(k))
    def cogs(k): return f"ABS({isc('costOfRevenueDisplayHeader', ycol(k))})"
    def gp(k): return isc("grossProfit", ycol(k))
    def opinc(k): return isc("operatingIncome", ycol(k))
    def ni(k): return isc("netIncome", ycol(k))
    def intexp(k): return f"ABS({isc('financeCosts', ycol(k))})"
    def ebitda(k): return isc("ebitda", ycol(k))
    def cash(k): return bsc("cashAndEquivalents", ycol(k))
    def ca(k): return bsc("currentAssetsHeader", ycol(k))
    def cl(k): return bsc("currentLiabilitiesHeader", ycol(k))
    def inv(k): return bsc("inventoryHeader", ycol(k))
    def ar(k): return bsc("receivablesHeader", ycol(k))
    def ap(k): return bsc("tradePayablesHeader", ycol(k))
    def ta(k): return bsc("assetsHeader", ycol(k))
    def tl(k): return bsc("liabilitiesHeader", ycol(k))
    def te(k): return bsc("equityHeader", ycol(k))
    def std(k): return f"({bsc('stBorrowingsData', ycol(k))}+{bsc('currentPortionLTDebt', ycol(k))})"
    def ltd(k): return bsc("ltDebtData", ycol(k))
    def tdebt(k): return f"({std(k)}+{ltd(k)})"
    def ocf(k): return cfc("operatingActivitiesHeader", ycol(k))

    add("LIQUIDITY", "Current Ratio", FMT_RATIO, lambda k: f"({ca(k)})/({cl(k)})")
    add("LIQUIDITY", "Quick Ratio", FMT_RATIO, lambda k: f"(({ca(k)})-({inv(k)}))/({cl(k)})")
    add("LIQUIDITY", "Cash Ratio", FMT_RATIO, lambda k: f"({cash(k)})/({cl(k)})")
    add("SOLVENCY", "Debt to Equity", FMT_RATIO, lambda k: f"({tl(k)})/({te(k)})")
    add("SOLVENCY", "Debt to Assets", FMT_RATIO, lambda k: f"{tdebt(k)}/({ta(k)})")
    add("SOLVENCY", "Debt to Capital", FMT_RATIO, lambda k: f"{tdebt(k)}/({tdebt(k)}+({te(k)}))")
    add("COVERAGE", "Interest Coverage", FMT_RATIO, lambda k: f"({opinc(k)})/({intexp(k)})")
    add("COVERAGE", "Debt Service Coverage", FMT_RATIO, lambda k: f"({ocf(k)})/({tdebt(k)}+({intexp(k)}))")
    add("COVERAGE", "Cash Flow to Debt", FMT_RATIO, lambda k: f"({ocf(k)})/{tdebt(k)}")
    add("PROFITABILITY", "Gross Margin", FMT_PCT, lambda k: f"({gp(k)})/({rev(k)})")
    add("PROFITABILITY", "Operating Margin", FMT_PCT, lambda k: f"({opinc(k)})/({rev(k)})")
    add("PROFITABILITY", "Net Profit Margin", FMT_PCT, lambda k: f"({ni(k)})/({rev(k)})")
    add("PROFITABILITY", "Return on Assets (ROA)", FMT_PCT, lambda k: f"({ni(k)})/({ta(k)})")
    add("PROFITABILITY", "Return on Equity (ROE)", FMT_PCT, lambda k: f"({ni(k)})/({te(k)})")
    add("PROFITABILITY", "Return on Invested Capital (ROIC)", FMT_PCT,
        lambda k: f"(({opinc(k)})*(1-({taxrate(ycol(k))})))/({tdebt(k)}+({te(k)})-({cash(k)}))")
    add("EFFICIENCY", "Asset Turnover", FMT_RATIO, lambda k: f"({rev(k)})/({avg(bs_l,'assetsHeader',k)})")
    add("EFFICIENCY", "Inventory Turnover", FMT_RATIO, lambda k: f"{cogs(k)}/({avg(bs_l,'inventoryHeader',k)})")
    add("EFFICIENCY", "Days Sales of Inventory", FMT_DAYS, lambda k: f"365/({cogs(k)}/({avg(bs_l,'inventoryHeader',k)}))")
    add("EFFICIENCY", "Receivables Turnover", FMT_RATIO, lambda k: f"({rev(k)})/({avg(bs_l,'receivablesHeader',k)})")
    add("EFFICIENCY", "Days Sales Outstanding", FMT_DAYS, lambda k: f"365/(({rev(k)})/({avg(bs_l,'receivablesHeader',k)}))")
    add("EFFICIENCY", "Accounts Payable Turnover", FMT_RATIO, lambda k: f"{cogs(k)}/({avg(bs_l,'tradePayablesHeader',k)})")
    add("EFFICIENCY", "Days Payable Outstanding", FMT_DAYS, lambda k: f"365/({cogs(k)}/({avg(bs_l,'tradePayablesHeader',k)}))")
    add("EFFICIENCY", "Cash Conversion Cycle", FMT_DAYS,
        lambda k: (f"365/({cogs(k)}/({avg(bs_l,'inventoryHeader',k)}))"
                   f"+365/(({rev(k)})/({avg(bs_l,'receivablesHeader',k)}))"
                   f"-365/({cogs(k)}/({avg(bs_l,'tradePayablesHeader',k)}))"))
    add("EFFICIENCY", "Working Capital Ratio", FMT_PCT, lambda k: f"(({ca(k)})-({cl(k)}))/({rev(k)})")

    # write
    r = _DATA_ROW = 8
    last_cat = None
    for cat, label, fmt, fn in rows:
        if cat != last_cat:
            cc = ws.cell(row=r, column=LABEL_COL, value=cat)
            cc.font = _font(bold=True, color=NAVY, size=10)
            for cidx in range(LABEL_COL, YEAR_COL0 + ncols):
                ws.cell(row=r, column=cidx).fill = PatternFill("solid", fgColor=GREY_HEADER)
            r += 1
            last_cat = cat
        ws.cell(row=r, column=LABEL_COL, value=label).font = _font(size=10, color="333333")
        ws.cell(row=r, column=LABEL_COL).alignment = Alignment(indent=1)
        for k in range(ncols):
            col = YEAR_COL0 + k
            cell = ws.cell(row=r, column=col)
            cell.value = f"=IFERROR({fn(k)},\"\")"
            cell.number_format = fmt
            cell.font = _font(color=GREEN_LINK, size=10)
        r += 1


def _ratio_year_header(ws, years, layout):
    hist = [y for y in years if y in layout.hist_years]
    proj = [y for y in years if y in layout.proj_years]
    hn = len(hist)
    ws.cell(row=7, column=LABEL_COL, value="").fill = PatternFill("solid", fgColor=NAVY)
    for i, y in enumerate(years):
        is_proj = y in layout.proj_years
        cc = ws.cell(row=7, column=YEAR_COL0 + i, value=f"{_parse_year(y)}{'P' if is_proj else 'A'}")
        cc.font = _font(bold=True, color=WHITE)
        cc.fill = PatternFill("solid", fgColor=NAVY_MED if is_proj else NAVY)
        cc.alignment = Alignment(horizontal="center")


# ============================================================
# HORIZONTAL ANALYSIS SHEET  (YoY % change, live)
# ============================================================
def _build_horizontal(ws, company, layouts):
    years = layouts[0].years
    trans = [(years[i - 1], years[i]) for i in range(1, len(years))]
    ncols = len(trans)
    _apply_col_widths(ws, ncols, label_width=40)
    ws.sheet_view.showGridLines = False
    _title_block(ws, f"{company} — HORIZONTAL ANALYSIS", ncols=ncols,
                 subtitle="Year-over-year % change · live formulas linked to the statements")
    # header
    ws.cell(row=5, column=LABEL_COL, value="Line Item").font = _font(bold=True, color=WHITE)
    ws.cell(row=5, column=LABEL_COL).fill = PatternFill("solid", fgColor=NAVY)
    for i, (p, c) in enumerate(trans):
        cc = ws.cell(row=5, column=YEAR_COL0 + i, value=f"{_parse_year(p)}→{_parse_year(c)}")
        cc.font = _font(bold=True, color=WHITE, size=9)
        cc.fill = PatternFill("solid", fgColor=NAVY_MED if c in layouts[0].proj_years else NAVY)
        cc.alignment = Alignment(horizontal="center")

    r = 6
    for layout in layouts:
        # statement band
        cc = ws.cell(row=r, column=LABEL_COL, value=layout.title.split("—")[-1].strip() or layout.sheet_name)
        cc.font = _font(bold=True, color=WHITE, size=10)
        for cidx in range(LABEL_COL, YEAR_COL0 + ncols):
            ws.cell(row=r, column=cidx).fill = PatternFill("solid", fgColor=NAVY_MED)
        cc.value = layout.sheet_name.upper()
        r += 1
        for tr in layout.template_rows:
            key = tr["key"]
            if tr.get("is_header"):
                continue
            ws.cell(row=r, column=LABEL_COL, value=tr["label"]).font = _font(size=10, color="333333")
            ws.cell(row=r, column=LABEL_COL).alignment = Alignment(indent=max(0, tr.get("level", 3) - 1))
            for i, (py, cy) in enumerate(trans):
                pcol = layout.year_col[py]
                ccol = layout.year_col[cy]
                prev = layout.cell(key, pcol)
                cur = layout.cell(key, ccol)
                cell = ws.cell(row=r, column=YEAR_COL0 + i)
                cell.value = f"=IFERROR(({cur}-{prev})/ABS({prev}),\"\")"
                cell.number_format = FMT_PCT
                cell.font = _font(color=GREEN_LINK, size=10)
            r += 1


# ============================================================
# DCF SHEET
# ============================================================
def _build_dcf(ws, company, is_l, bs_l, cf_l, dcf_row, n_forecast):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 34
    for col in "CDEFGHIJKL":
        ws.column_dimensions[col].width = 12
    _title_block(ws, f"{company} — DCF VALUATION", ncols=8,
                 subtitle="Discounted Cash Flow · live model · blue cells are editable assumptions")

    base_col = is_l.base_col
    hist = is_l.hist_years
    prev_col = is_l.year_col[hist[-2]] if len(hist) >= 2 else None

    def isc(key, col=base_col): return is_l.cell(key, col)
    def bsc(key, col=base_col): return bs_l.cell(key, col)
    def cfc(key, col=base_col): return cf_l.cell(key, col)

    # ── Key assumptions — live links to the Assumptions sheet ──
    # WACC/TGR/method/exit multiple/shares are edited on the Assumptions sheet
    # (merged there alongside the forecast drivers); the DCF sheet just
    # displays them as cross-sheet formulas, same convention as BASE METRICS.
    R = {}
    r = 5
    ws.cell(row=r, column=2, value="KEY ASSUMPTIONS  (edit on the Assumptions sheet)").font = _font(bold=True, color=NAVY)
    for cidx in range(2, 5):
        ws.cell(row=r, column=cidx).fill = PatternFill("solid", fgColor=GREY_HEADER)
    r += 1

    ASq = _q("Assumptions")

    def link(label, key, dcf_name, fmt=FMT_NUM1):
        nonlocal r
        ws.cell(row=r, column=2, value=label).font = _font(size=10, color="333333")
        c = ws.cell(row=r, column=3, value=f"={ASq}!C{dcf_row[dcf_name]}")
        c.font = _font(color=GREEN_LINK, bold=True)
        c.number_format = fmt
        R[key] = f"C{r}"
        r += 1

    link("WACC (%)", "wacc", "wacc")
    link("Terminal Growth Rate (%)", "tgr", "tgr")
    link("Method", "method", "method", fmt="General")
    link("Exit Multiple (x)", "exit", "exit_mult")
    link("Shares Outstanding", "shares", "shares", fmt=FMT_NUM)
    link("Forecast Years", "yrs", "fcst_years", fmt="0")

    # ── Base metrics block (formulas from statements) ──
    r += 1
    ws.cell(row=r, column=2, value="BASE METRICS (from latest actual)").font = _font(bold=True, color=NAVY)
    for cidx in range(2, 5):
        ws.cell(row=r, column=cidx).fill = PatternFill("solid", fgColor=GREY_HEADER)
    r += 1

    def metric(label, key, formula, fmt=FMT_NUM):
        nonlocal r
        ws.cell(row=r, column=2, value=label).font = _font(size=10, color="333333")
        c = ws.cell(row=r, column=3)
        c.value = "=" + formula
        c.number_format = fmt
        c.font = _font(color=GREEN_LINK)
        R[key] = f"C{r}"
        r += 1

    dep = f"ABS({isc('depreciationCostOfSales')}+{isc('depreciationOpex')})"
    ebt = isc("earningsBeforeTax")
    tax = f"ABS({isc('incomeTaxExpense')})"
    # Nested IF so tax/ebt is only evaluated when ebt>0 (avoids #DIV/0! that a
    # single AND(...) would trigger by evaluating every argument).
    taxrate = f"IF({ebt}>0,IF(AND({tax}/{ebt}>=0,{tax}/{ebt}<=1),{tax}/{ebt},0.25),0.25)"
    opinc = isc("operatingIncome")
    capex = f"ABS({cfc('capitalExpenditures')})"
    if prev_col:
        ca_b = bsc("currentAssetsHeader"); ca_p = bs_l.cell("currentAssetsHeader", prev_col)
        cl_b = bsc("currentLiabilitiesHeader"); cl_p = bs_l.cell("currentLiabilitiesHeader", prev_col)
        wc = f"(({ca_b}-{ca_p})-({cl_b}-{cl_p}))"
    else:
        wc = "0"
    curr_debt = f"({bsc('stBorrowingsData')}+{bsc('currentPortionLTDebt')})"
    tdebt = f"({curr_debt}+{bsc('ltDebtData')})"

    metric("Tax Rate", "taxrate", taxrate, FMT_PCT)
    metric("EBITDA", "ebitda", f"{opinc}+{dep}")
    metric("Net Debt", "netdebt", f"{tdebt}-{bsc('cashAndEquivalents')}")
    metric("NOPAT", "nopat", f"{opinc}*(1-{R['taxrate']})")
    metric("Base Free Cash Flow", "basefcf", f"{R['nopat']}+{dep}-{capex}-{wc}")

    # ── FCF build-up ──
    r += 1
    ws.cell(row=r, column=2, value="DCF PROJECTION").font = _font(bold=True, color=NAVY)
    for cidx in range(2, 3 + n_forecast + 1):
        ws.cell(row=r, column=cidx).fill = PatternFill("solid", fgColor=GREY_HEADER)
    hdr_row = r
    r += 1
    # year labels
    ws.cell(row=r, column=2, value="Forecast Year").font = _font(bold=True, size=9, color="555555")
    for k in range(n_forecast):
        ws.cell(row=r, column=3 + k, value=f"Y{k+1}").font = _font(bold=True, color=WHITE)
        ws.cell(row=r, column=3 + k).fill = PatternFill("solid", fgColor=NAVY_MED)
        ws.cell(row=r, column=3 + k).alignment = Alignment(horizontal="center")
    yr_row = r
    r += 1
    # fcf growth rate
    fcfg = f"MIN({R['tgr']}/100*1.5,0.08)"
    ws.cell(row=r, column=2, value="FCF Growth Rate").font = _font(size=10, color="333333")
    ws.cell(row=r, column=3, value=f"={fcfg}").number_format = FMT_PCT
    ws.cell(row=r, column=3).font = _font(color=GREEN_LINK)
    fcfg_cell = f"$C${r}"
    r += 1
    # FCF row
    ws.cell(row=r, column=2, value="Projected FCF").font = _font(size=10, color="333333")
    fcf_row = r
    for k in range(n_forecast):
        col = get_column_letter(3 + k)
        if k == 0:
            f = f"={R['basefcf']}*(1+{fcfg_cell})"
        else:
            prevc = get_column_letter(2 + k)
            f = f"={prevc}{fcf_row}*(1+{fcfg_cell})"
        c = ws.cell(row=r, column=3 + k, value=f)
        c.number_format = FMT_NUM; c.font = _font(color=BLACK)
    r += 1
    # discount factor
    ws.cell(row=r, column=2, value="Discount Factor").font = _font(size=10, color="333333")
    df_row = r
    for k in range(n_forecast):
        c = ws.cell(row=r, column=3 + k, value=f"=1/(1+{R['wacc']}/100)^{k+1}")
        c.number_format = '0.000'; c.font = _font(color=BLACK)
    r += 1
    # PV of FCF
    ws.cell(row=r, column=2, value="PV of FCF").font = _font(size=10, color="333333")
    pv_row = r
    for k in range(n_forecast):
        col = get_column_letter(3 + k)
        c = ws.cell(row=r, column=3 + k, value=f"={col}{fcf_row}*{col}{df_row}")
        c.number_format = FMT_NUM; c.font = _font(color=BLACK, bold=True)
    r += 2

    # ── Valuation summary ──
    lastfcf = f"{get_column_letter(3 + n_forecast - 1)}{fcf_row}"
    sumpv = f"SUM(C{pv_row}:{get_column_letter(3 + n_forecast - 1)}{pv_row})"

    def summ(label, key, formula, fmt=FMT_NUM, bold=False, big=False):
        nonlocal r
        lc = ws.cell(row=r, column=2, value=label)
        lc.font = _font(bold=bold or big, color=NAVY if big else "333333", size=11 if big else 10)
        c = ws.cell(row=r, column=3)
        c.value = "=" + formula
        c.number_format = fmt
        c.font = _font(color=BLACK, bold=bold or big, size=11 if big else 10)
        if big:
            c.fill = PatternFill("solid", fgColor=ACCENT_PROJ)
        R[key] = f"C{r}"
        r += 1

    summ("Sum of PV of FCF", "sumpv", sumpv)
    # terminal value: perpetuity vs multiple. IFERROR guards WACC<=TGR (Gordon
    # growth undefined) so the workbook never shows #DIV/0!.
    tv_perp = f"{lastfcf}*(1+{R['tgr']}/100)/(({R['wacc']}-{R['tgr']})/100)"
    tv_mult = f"{R['ebitda']}*(1+{fcfg_cell})^{n_forecast}*{R['exit']}"
    summ("Terminal Value", "tv", f"IFERROR(IF({R['method']}=\"Perpetuity\",{tv_perp},{tv_mult}),0)")
    summ("PV of Terminal Value", "pvtv", f"{R['tv']}/(1+{R['wacc']}/100)^{n_forecast}")
    summ("Enterprise Value", "ev", f"{R['sumpv']}+{R['pvtv']}", bold=True)
    summ("Less: Net Debt", "less_nd", f"-{R['netdebt']}")
    summ("Equity Value", "equity", f"{R['ev']}-{R['netdebt']}", bold=True)
    summ("Value per Share", "vps", f"IF({R['shares']}>0,{R['equity']}/{R['shares']},\"\")",
         fmt='#,##0.00', big=True)

    # ── Sensitivity grid (WACC × TGR, perpetuity) ──
    r += 2
    ws.cell(row=r, column=2, value="SENSITIVITY — Value per Share (WACC × Terminal Growth)").font = _font(bold=True, color=NAVY)
    r += 1
    grid_top = r
    # WACC axis: base ±2 step 0.5 → 9 cols starting col C
    wacc_offsets = [i * 0.5 for i in range(-4, 5)]
    tgr_offsets = [i * 0.25 for i in range(-4, 5)]
    ws.cell(row=r, column=2, value="TGR ↓ / WACC →").font = _font(bold=True, size=9, color="555555")
    for j, wo in enumerate(wacc_offsets):
        c = ws.cell(row=r, column=3 + j, value=f"={R['wacc']}+{wo}")
        c.number_format = '0.0'; c.font = _font(bold=True, color=WHITE); c.fill = PatternFill("solid", fgColor=NAVY_MED)
        c.alignment = Alignment(horizontal="center")
    wacc_hdr_row = r
    r += 1
    for i, to in enumerate(tgr_offsets):
        rc = ws.cell(row=r, column=2, value=f"={R['tgr']}+{to}")
        rc.number_format = '0.00'; rc.font = _font(bold=True, color=WHITE); rc.fill = PatternFill("solid", fgColor=NAVY_MED)
        for j, wo in enumerate(wacc_offsets):
            wcell = f"{get_column_letter(3 + j)}${wacc_hdr_row}"
            tcell = f"$B{r}"
            g = f"MIN({tcell}/100*1.5,0.08)"
            # totalPV = basefcf*SUMPRODUCT(((1+g)/(1+w))^{1..5})
            powers = ",".join(str(x) for x in range(1, n_forecast + 1))
            totalpv = f"{R['basefcf']}*SUMPRODUCT(((1+{g})/(1+{wcell}/100))^{{{powers}}})"
            fcfN = f"{R['basefcf']}*(1+{g})^{n_forecast}"
            tv = f"{fcfN}*(1+{tcell}/100)/(({wcell}-{tcell})/100)"
            pvtv = f"{tv}/(1+{wcell}/100)^{n_forecast}"
            eq = f"(({totalpv})+({pvtv})-{R['netdebt']})"
            cell = ws.cell(row=r, column=3 + j)
            cell.value = f"=IF({tcell}>={wcell}-0.25,\"\",IFERROR(IF({R['shares']}>0,{eq}/{R['shares']},\"\"),\"\"))"
            cell.number_format = '#,##0.00'
            cell.font = _font(size=9)
            if i == 4 and j == 4:
                cell.fill = PatternFill("solid", fgColor="FFFF00")
        r += 1


# ============================================================
# ORCHESTRATION
# ============================================================
def build_workbook(project: dict) -> Workbook:
    company = (project.get("company_name") or "Company").strip()
    currency = project.get("currency") or "SAR"
    templates = load_statement_templates()

    is_stored = project.get("income_statement")
    bs_stored = project.get("balance_sheet")
    cf_stored = project.get("cash_flow_statement")

    def yrs(s):
        s = s or {}
        years = list(s.get("years") or [])
        if years:
            return years
        # Fallback: derive from the row value keys when "years" isn't stored.
        found = set()
        for row in s.get("rows") or []:
            found.update((row.get("values") or {}).keys())
        return list(found)

    hist = sorted(set(yrs(is_stored) + yrs(bs_stored) + yrs(cf_stored)), key=_parse_year)
    if not hist:
        raise ValueError("Project has no financial statement data to export.")
    base_year = hist[-1]
    base_year_int = _parse_year(base_year)

    fd = project.get("forecast_data") or {}
    n_forecast = int(fd.get("forecast_years") or 5)
    # Default the projection basis to FAITHFUL so an unbalanced historical
    # balance sheet surfaces in the Balance Check row every year (an analyst
    # can see and diagnose it) rather than being hidden by the cash plug.
    # A saved forecast's own mode is respected when present.
    balance_mode = fd.get("balance_mode") or "faithful"
    mode_label = ("Faithful (carries BS imbalance)" if balance_mode == "faithful"
                  else "Balanced (cash plug)")
    inputs = fd.get("inputs")
    if not inputs:
        inputs = asdict(calculate_historical_assumptions(is_stored, bs_stored))

    proj_years = [str(base_year_int + i) for i in range(1, n_forecast + 1)]

    def sub(title_kind):
        first = _parse_year(hist[0])
        return (f"Amounts in {currency} '000  ·  Actual FY{first}–FY{base_year_int}  ·  "
                f"Projected FY{base_year_int + 1}–FY{base_year_int + n_forecast}  ·  "
                f"Projection basis: {mode_label}  ·  Confidential")

    # Layouts
    is_l = StmtLayout("Income Statement", f"{company}  —  INCOME STATEMENT",
                      templates["income_statement"], is_stored, hist, proj_years)
    bs_l = StmtLayout("Balance Sheet", f"{company}  —  BALANCE SHEET",
                      templates["balance_sheet"], bs_stored, hist, proj_years)
    cf_l = StmtLayout("Cash Flow Statement", f"{company}  —  CASH FLOW STATEMENT",
                      templates["cash_flow_statement"], cf_stored, hist, proj_years)

    # Workbook + sheets (display order)
    wb = Workbook()
    ws_is = wb.active
    ws_is.title = "Income Statement"
    ws_bs = wb.create_sheet("Balance Sheet")
    ws_cf = wb.create_sheet("Cash Flow Statement")
    ws_ratios = wb.create_sheet("Ratios")
    ws_horiz = wb.create_sheet("Horizontal Analysis")
    ws_dcf = wb.create_sheet("DCF")
    ws_assum = wb.create_sheet(MODEL_SHEET)

    # Base-year anchors, resolved exactly as the forecasting engine does.
    base = extract_base_data(is_stored, bs_stored, cf_stored,
                             dcf_assumptions=project.get("dcf_assumptions"))

    # Build in dependency order — the driver panel first, then the engine
    # calculation block appended below it on the same sheet.
    assum_row, dcf_row, banner_row = _build_assumptions(
        ws_assum, company, base_year_int, proj_years, inputs,
        base, is_l, bs_l, cf_l, balance_mode,
        mode_label, dcf_assumptions=project.get("dcf_assumptions"))
    first_var_row = banner_row + 2   # banner, block header, then variables
    engine_row = {name: first_var_row + i for i, (name, _, _) in enumerate(ENGINE_VARS)}
    _build_engine(ws_assum, engine_row, base, assum_row, n_forecast, balance_mode,
                  banner_row, base_year_int, proj_years)

    overrides = _build_overrides(engine_row)
    _build_statement(ws_is, is_l, sub("is"), overrides["income_statement"], engine_row,
                     "rev_mult", "income_statement")
    _build_statement(ws_bs, bs_l, sub("bs"), overrides["balance_sheet"], engine_row,
                     "ast_mult", "balance_sheet")
    _build_statement(ws_cf, cf_l, sub("cf"), overrides["cash_flow_statement"], engine_row,
                     "rev_mult", "cash_flow_statement")

    _build_ratios(ws_ratios, company, is_l, bs_l, cf_l)
    _build_horizontal(ws_horiz, company, [is_l, bs_l, cf_l])
    _build_dcf(ws_dcf, company, is_l, bs_l, cf_l, dcf_row, n_forecast)

    return wb


def export_project_to_xlsx_bytes(project: dict) -> bytes:
    wb = build_workbook(project)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
