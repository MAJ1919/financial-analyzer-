"""
The shipped download template must stay in sync with the canonical template.

frontend/public/Saudi_Template.xlsx is what users download, fill in and
re-upload. excel_parser.py matches its rows against manualEntryTemplate.json
STRICTLY, so any label drift between the two makes the user's own download
come back with "Unmapped Rows Detected".

Regenerate with:  python scripts/generate_templates.py
"""
from pathlib import Path

import openpyxl
import pytest

from app.services.excel_parser import (
    STRICT_SHEET_MAP,
    TEMPLATE_DATA,
    parse_template_upload,
)

TEMPLATE_XLSX = (
    Path(__file__).resolve().parents[2] / "frontend" / "public" / "Saudi_Template.xlsx"
)

SHEET_TO_STMT = {
    "Income Statement": "income_statement",
    "Balance Sheet": "balance_sheet",
    "Cash Flow Statement": "cash_flow_statement",
}


@pytest.fixture(scope="module")
def workbook_bytes() -> bytes:
    assert TEMPLATE_XLSX.exists(), f"shipped template missing: {TEMPLATE_XLSX}"
    return TEMPLATE_XLSX.read_bytes()


def sheet_labels(wb, sheet_name: str) -> list[str]:
    """Data-row labels (column A, below the header row), stripped of indentation."""
    ws = wb[sheet_name]
    labels = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and isinstance(row[0], str) and row[0].strip():
            labels.append(row[0].strip())
    return labels


class TestShippedTemplateWorkbook:
    def test_sheet_names_are_strictly_recognised(self, workbook_bytes):
        wb = openpyxl.load_workbook(TEMPLATE_XLSX)
        assert set(wb.sheetnames) == set(SHEET_TO_STMT)
        for name in wb.sheetnames:
            assert name.lower().strip() in STRICT_SHEET_MAP

    @pytest.mark.parametrize("sheet_name, stmt_type", SHEET_TO_STMT.items())
    def test_labels_match_canonical_template_exactly(self, workbook_bytes, sheet_name, stmt_type):
        wb = openpyxl.load_workbook(TEMPLATE_XLSX)
        shipped = sheet_labels(wb, sheet_name)
        canonical = [r["label"].strip() for r in TEMPLATE_DATA[stmt_type]]

        missing = [lbl for lbl in canonical if lbl not in shipped]
        extra = [lbl for lbl in shipped if lbl not in canonical]
        assert not missing, f"{sheet_name}: missing from workbook {missing}"
        assert not extra, f"{sheet_name}: not in canonical template {extra}"
        assert shipped == canonical, f"{sheet_name}: row order drifted"

    def test_upload_of_the_shipped_template_has_zero_unmapped_rows(self, workbook_bytes):
        """The exact symptom this guards: downloading the template, then
        re-uploading it, must not produce an 'Unmapped Rows Detected' warning."""
        result = parse_template_upload(workbook_bytes)
        assert result.get("unmapped_rows") == {}

    def test_every_template_row_is_addressable_after_upload(self, workbook_bytes):
        result = parse_template_upload(workbook_bytes)
        for sheet_name, stmt_type in SHEET_TO_STMT.items():
            parsed_keys = [r["key"] for r in result[stmt_type]["rows"]]
            canonical_keys = [r["key"] for r in TEMPLATE_DATA[stmt_type]]
            assert parsed_keys == canonical_keys, sheet_name

    def test_year_headers_are_parseable(self, workbook_bytes):
        """A workbook without recognisable year columns raises on upload."""
        result = parse_template_upload(workbook_bytes)
        years = result["income_statement"]["years"]
        assert years, "no year columns found in the shipped template"
        assert years == sorted(years, key=int)
