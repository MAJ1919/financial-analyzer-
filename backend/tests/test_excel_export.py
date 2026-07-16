"""
Tests for the Excel export service (app.services.excel_export).

These assert the workbook's STRUCTURE and that projections/analytics are wired
as live formulas — the numeric fidelity to the forecasting engine is validated
separately. Kept dependency-free (openpyxl only; no LibreOffice/recalc).
"""
import io

import openpyxl
import pytest

from app.services.excel_export import build_workbook, export_project_to_xlsx_bytes
from app.services.forecasting_engine import run_forecast, calculate_historical_assumptions
from dataclasses import asdict


def _cash_flow() -> dict:
    return {
        "rows": [
            {"row_id": "operatingActivitiesHeader", "key": "operatingActivitiesHeader",
             "label": "Operating Activities", "section": "Operating", "is_header": True,
             "values": {"2022": 1500.0, "2023": 1700.0}},
            {"row_id": "capitalExpenditures", "key": "capitalExpenditures",
             "label": "Capital Expenditures", "section": "Investing",
             "values": {"2022": -400.0, "2023": -450.0}},
        ]
    }


@pytest.fixture
def project(income_statement, balance_sheet) -> dict:
    inputs = asdict(calculate_historical_assumptions(income_statement, balance_sheet))
    fc = run_forecast(
        income_statement=income_statement,
        balance_sheet=balance_sheet,
        cash_flow=_cash_flow(),
        inputs=inputs,
        scenarios=["base"],
        forecast_years=5,
        balance_mode="balanced",
    )
    return {
        "company_name": "Test Co",
        "currency": "SAR",
        "income_statement": income_statement,
        "balance_sheet": balance_sheet,
        "cash_flow_statement": _cash_flow(),
        "forecast_data": fc,
        "dcf_assumptions": {"wacc": 10, "terminal_growth_rate": 2.5, "shares_outstanding": 1000},
    }


def _find_row(ws, label, col="B", start=1):
    for r in range(start, ws.max_row + 1):
        if ws[f"{col}{r}"].value == label:
            return r
    return None


def test_workbook_has_all_sheets_engine_merged(project):
    """The forecast engine lives ON the Assumptions sheet (grouped rows under
    a banner), not on a separate hidden tab."""
    wb = build_workbook(project)
    assert wb.sheetnames == [
        "Income Statement", "Balance Sheet", "Cash Flow Statement",
        "Ratios", "Horizontal Analysis", "DCF", "Assumptions",
    ]
    ws = wb["Assumptions"]
    banner = _find_row(ws, "FORECAST ENGINE — CALCULATED (do not edit; driven by the drivers above)")
    assert banner is not None
    # the calculation block is grouped one outline level deep and collapsed
    first_var = banner + 2
    assert ws.row_dimensions[first_var].outline_level == 1
    assert ws.row_dimensions[first_var].hidden is True
    assert ws.row_dimensions[banner].collapsed is True


def test_projected_revenue_links_to_engine(project):
    wb = build_workbook(project)
    ws = wb["Income Statement"]
    r = _find_row(ws, "Revenue")
    assert r is not None
    # hist years 2022,2023 → first projected column is E (C=2022, D=2023, E=2024P)
    val = ws[f"E{r}"].value
    assert isinstance(val, str) and val.startswith("=")
    assert "Assumptions" in val


def test_historical_actual_is_stored_value(project):
    wb = build_workbook(project)
    ws = wb["Income Statement"]
    # Income Tax Expense uses a template key the fixture populates.
    r = _find_row(ws, "Income Tax Expense")
    assert r is not None
    # 2023 (base actual) is column D and holds the stored number, not a formula
    assert ws[f"D{r}"].value == 400.0


def test_ratios_and_horizontal_are_live_formulas(project):
    wb = build_workbook(project)
    ratios = wb["Ratios"]
    cr = _find_row(ratios, "Current Ratio")
    assert cr is not None
    formula = ratios[f"C{cr}"].value
    assert isinstance(formula, str) and formula.startswith("=IFERROR")

    horiz = wb["Horizontal Analysis"]
    # first data formula cell should be a YoY IFERROR percentage
    found = any(
        isinstance(horiz[f"C{r}"].value, str) and str(horiz[f"C{r}"].value).startswith("=IFERROR")
        for r in range(6, horiz.max_row + 1)
    )
    assert found


def test_dcf_has_value_per_share_formula(project):
    wb = build_workbook(project)
    dcf = wb["DCF"]
    r = _find_row(dcf, "Value per Share")
    assert r is not None
    assert isinstance(dcf[f"C{r}"].value, str) and dcf[f"C{r}"].value.startswith("=")


def test_export_bytes_are_valid_xlsx(project):
    data = export_project_to_xlsx_bytes(project)
    assert data[:2] == b"PK"  # xlsx is a zip
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert "DCF" in wb.sheetnames


def test_export_without_data_raises():
    with pytest.raises(ValueError):
        build_workbook({"company_name": "Empty"})


def test_export_works_without_forecast_or_dcf(income_statement, balance_sheet):
    """Export must succeed (using derived assumptions) even if the user has
    not run a forecast or DCF yet."""
    proj = {
        "company_name": "No Forecast Co",
        "income_statement": income_statement,
        "balance_sheet": balance_sheet,
    }
    wb = build_workbook(proj)
    assert "Assumptions" in wb.sheetnames
