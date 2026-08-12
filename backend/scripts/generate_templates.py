"""
Regenerate the blank Excel upload template shipped to users.

The workbook is generated FROM the canonical template
(app/models/manualEntryTemplate.json) so the labels it contains are, by
construction, the labels excel_parser.py matches against. Any other source
lets the download drift out of sync and every drifted row lands in
`unmapped_rows` on upload.

Run from backend/:
    python scripts/generate_templates.py

Guarded by tests/test_template_workbook.py, which fails if the committed
workbook stops matching the canonical template.
"""
import datetime as _dt
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

BACKEND_DIR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(BACKEND_DIR))

from app.models.statement_templates import load_statement_templates  # noqa: E402

OUTPUT_PATH = BACKEND_DIR.parent / "frontend" / "public" / "Saudi_Template.xlsx"

# Sheet titles must match excel_parser.STRICT_SHEET_MAP exactly.
SHEET_ORDER = [
    ("Income Statement", "income_statement"),
    ("Balance Sheet", "balance_sheet"),
    ("Cash Flow Statement", "cash_flow_statement"),
]

# Placeholder year columns: the four most recent completed fiscal years at
# generation time, rather than a hardcoded window that goes stale on the shelf.
# The parser reads whatever years the header row holds, so users may overwrite
# or extend these.
_LAST_COMPLETE_YEAR = _dt.date.today().year - 1
YEARS = [str(y) for y in range(_LAST_COMPLETE_YEAR - 3, _LAST_COMPLETE_YEAR + 1)]


def build_workbook(templates: dict) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name, stmt_type in SHEET_ORDER:
        rows = templates.get(stmt_type, [])
        ws = wb.create_sheet(title=sheet_name)

        ws.cell(row=1, column=1, value="Line Item").font = Font(bold=True)
        for col, year in enumerate(YEARS, start=2):
            cell = ws.cell(row=1, column=col, value=year)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="right")

        for r_idx, t_row in enumerate(rows, start=2):
            label = (t_row.get("label") or "").strip()
            # Bloomberg-style indent: level 1 -> 0 spaces, level 2 -> 2, ...
            # excel_parser strips the label, so indentation is display-only.
            indent = " " * ((int(t_row.get("level", 3)) - 1) * 2)
            cell = ws.cell(row=r_idx, column=1, value=indent + label)
            if t_row.get("is_header") or t_row.get("is_subtotal"):
                cell.font = Font(bold=True)

        ws.column_dimensions["A"].width = 50
        for col_letter in ("B", "C", "D", "E"):
            ws.column_dimensions[col_letter].width = 15
        ws.freeze_panes = "B2"

    return wb


if __name__ == "__main__":
    templates = load_statement_templates()
    wb = build_workbook(templates)
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)

    total = sum(len(templates.get(t, [])) for _, t in SHEET_ORDER)
    print(f"Wrote {OUTPUT_PATH} ({total} line items across {len(SHEET_ORDER)} sheets)")
