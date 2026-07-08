import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

def parse_markdown(filepath):
    with open(filepath, 'r', encoding='utf-8') as f:
        lines = f.readlines()
        
    sheets = {}
    current_sheet = None
    
    for line in lines:
        line = line.strip()
        if not line or line.startswith('----') or line.startswith('===='):
            continue
            
        if line == 'INCOME STATEMENT':
            current_sheet = 'Income Statement'
            sheets[current_sheet] = []
            continue
        elif line == 'BALANCE SHEET':
            current_sheet = 'Balance Sheet'
            sheets[current_sheet] = []
            continue
        elif line.startswith('CASH FLOW STATEMENT'):
            current_sheet = 'Cash Flow Statement'
            sheets[current_sheet] = []
            continue
            
        if current_sheet is None:
            continue
            
        if line.startswith('# '):
            # Level 1
            sheets[current_sheet].append({'label': line[2:].strip(), 'level': 1})
        elif line.startswith('## '):
            # Level 2
            sheets[current_sheet].append({'label': line[3:].strip(), 'level': 2})
        elif line.startswith('### '):
            # Level 3
            sheets[current_sheet].append({'label': line[4:].strip(), 'level': 3})
        elif line.startswith('- '):
            # Line item (Level 4, or whatever the parent level is + 1, let's just use spaces)
            # Actually bloomberg style:
            # level 1 = 0 spaces
            # level 2 = 2 spaces
            # level 3 = 4 spaces
            sheets[current_sheet].append({'label': line[2:].strip(), 'level': 4})
            
    return sheets

def create_excel(sheets, output_file):
    wb = Workbook()
    wb.remove(wb.active) # remove default sheet
    
    years = ['2021', '2022', '2023']
    
    for sheet_name, rows in sheets.items():
        ws = wb.create_sheet(title=sheet_name)
        
        # Headers
        ws.cell(row=1, column=1, value="Line Item").font = Font(bold=True)
        for col, year in enumerate(years, start=2):
            ws.cell(row=1, column=col, value=year).font = Font(bold=True)
            ws.cell(row=1, column=col).alignment = Alignment(horizontal='right')
            
        # Data
        for r_idx, row_data in enumerate(rows, start=2):
            level = row_data['level']
            label = row_data['label']
            
            # Add spaces based on level (bloomberg style)
            # Level 1 -> 0 spaces
            # Level 2 -> 2 spaces
            # Level 3 -> 4 spaces
            # Level 4 -> 6 spaces
            spaces = " " * ((level - 1) * 2)
            final_label = spaces + label
            
            cell = ws.cell(row=r_idx, column=1, value=final_label)
            if level <= 3:
                cell.font = Font(bold=True)
                
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        
    wb.save(output_file)

if __name__ == '__main__':
    # Saudi Core IFRS
    sheets = parse_markdown(r'c:\Users\s9378\Desktop\Summer Work\Code Work\Code\Saudi_Core_IFRS.md')
    create_excel(sheets, r'c:\Users\s9378\Desktop\Summer Work\Code Work\financial-analyzer-platform\frontend\public\Saudi_Template.xlsx')
    
    print("Excel template generated successfully in frontend/public/")
